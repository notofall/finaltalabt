"""
نظام إدارة كميات العمائر السكنية - Buildings System Routes
PostgreSQL Implementation
"""
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, update
from sqlalchemy.orm import selectinload
from typing import Optional, List
from datetime import datetime, timezone
import uuid
import json
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database.connection import get_postgres_session
from database.models import (
    User, Project, PriceCatalogItem, 
    UnitTemplate, UnitTemplateMaterial, ProjectFloor, 
    ProjectAreaMaterial, SupplyTracking
)
from routes.pg_auth_routes import get_current_user_pg

pg_buildings_router = APIRouter(prefix="/api/pg/buildings")


# ==================== نماذج الوحدات ====================

@pg_buildings_router.get("/projects/{project_id}/templates")
async def get_unit_templates(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """الحصول على نماذج الوحدات للمشروع"""
    result = await session.execute(
        select(UnitTemplate).where(UnitTemplate.project_id == project_id)
    )
    templates = result.scalars().all()
    
    templates_list = []
    for t in templates:
        # جلب مواد النموذج
        materials_result = await session.execute(
            select(UnitTemplateMaterial).where(UnitTemplateMaterial.template_id == t.id)
        )
        materials = materials_result.scalars().all()
        
        templates_list.append({
            "id": t.id,
            "code": t.code,
            "name": t.name,
            "description": t.description,
            "area": t.area,
            "rooms_count": t.rooms_count,
            "bathrooms_count": t.bathrooms_count,
            "count": t.count,
            "project_id": t.project_id,
            "materials": [{
                "id": m.id,
                "catalog_item_id": m.catalog_item_id,
                "item_code": m.item_code,
                "item_name": m.item_name,
                "unit": m.unit,
                "quantity_per_unit": m.quantity_per_unit,
                "unit_price": m.unit_price
            } for m in materials],
            "created_at": t.created_at.isoformat() if t.created_at else None
        })
    
    return templates_list


@pg_buildings_router.post("/projects/{project_id}/templates")
async def create_unit_template(
    project_id: str,
    template_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إنشاء نموذج وحدة جديد"""
    # التحقق من وجود المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    new_template = UnitTemplate(
        id=str(uuid.uuid4()),
        code=template_data.get("code", ""),
        name=template_data.get("name", ""),
        description=template_data.get("description"),
        area=template_data.get("area", 0),
        rooms_count=template_data.get("rooms_count", 0),
        bathrooms_count=template_data.get("bathrooms_count", 0),
        count=template_data.get("count", 0),
        project_id=project_id,
        project_name=project.name,
        created_by=current_user.id,
        created_by_name=current_user.name,
        created_at=datetime.now(timezone.utc).replace(tzinfo=None)
    )
    
    session.add(new_template)
    await session.commit()
    
    return {
        "id": new_template.id,
        "code": new_template.code,
        "name": new_template.name,
        "message": "تم إنشاء النموذج بنجاح"
    }


@pg_buildings_router.put("/projects/{project_id}/templates/{template_id}")
async def update_unit_template(
    project_id: str,
    template_id: str,
    template_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحديث نموذج وحدة"""
    result = await session.execute(
        select(UnitTemplate).where(
            UnitTemplate.id == template_id,
            UnitTemplate.project_id == project_id
        )
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="النموذج غير موجود")
    
    for key in ["code", "name", "description", "area", "rooms_count", "bathrooms_count", "count"]:
        if key in template_data:
            setattr(template, key, template_data[key])
    
    template.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    await session.commit()
    
    return {"message": "تم التحديث بنجاح"}


@pg_buildings_router.delete("/projects/{project_id}/templates/{template_id}")
async def delete_unit_template(
    project_id: str,
    template_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """حذف نموذج وحدة"""
    await session.execute(
        delete(UnitTemplateMaterial).where(UnitTemplateMaterial.template_id == template_id)
    )
    await session.execute(
        delete(UnitTemplate).where(
            UnitTemplate.id == template_id,
            UnitTemplate.project_id == project_id
        )
    )
    await session.commit()
    return {"message": "تم الحذف بنجاح"}


# ==================== مواد نموذج الوحدة ====================

@pg_buildings_router.post("/templates/{template_id}/materials")
async def add_template_material(
    template_id: str,
    material_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إضافة مادة لنموذج الوحدة"""
    # التحقق من وجود النموذج
    template_result = await session.execute(
        select(UnitTemplate).where(UnitTemplate.id == template_id)
    )
    template = template_result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="النموذج غير موجود")
    
    # جلب بيانات المادة من الكتالوج
    catalog_item_id = material_data.get("catalog_item_id")
    catalog_result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.id == catalog_item_id)
    )
    catalog_item = catalog_result.scalar_one_or_none()
    
    new_material = UnitTemplateMaterial(
        id=str(uuid.uuid4()),
        template_id=template_id,
        catalog_item_id=catalog_item_id,
        item_code=catalog_item.item_code if catalog_item else material_data.get("item_code"),
        item_name=catalog_item.name if catalog_item else material_data.get("item_name"),
        unit=catalog_item.unit if catalog_item else material_data.get("unit", "قطعة"),
        quantity_per_unit=material_data.get("quantity_per_unit", 0),
        unit_price=catalog_item.price if catalog_item else material_data.get("unit_price", 0)
    )
    
    session.add(new_material)
    await session.commit()
    
    return {"id": new_material.id, "message": "تمت الإضافة بنجاح"}


@pg_buildings_router.delete("/templates/{template_id}/materials/{material_id}")
async def remove_template_material(
    template_id: str,
    material_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """حذف مادة من نموذج الوحدة"""
    await session.execute(
        delete(UnitTemplateMaterial).where(
            UnitTemplateMaterial.id == material_id,
            UnitTemplateMaterial.template_id == template_id
        )
    )
    await session.commit()
    return {"message": "تم الحذف بنجاح"}


# ==================== أدوار المشروع ====================

@pg_buildings_router.get("/projects/{project_id}/floors")
async def get_project_floors(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """الحصول على أدوار المشروع"""
    result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id).order_by(ProjectFloor.floor_number)
    )
    floors = result.scalars().all()
    
    return [{
        "id": f.id,
        "floor_number": f.floor_number,
        "floor_name": f.floor_name,
        "area": f.area,
        "steel_factor": f.steel_factor
    } for f in floors]


@pg_buildings_router.post("/projects/{project_id}/floors")
async def add_project_floor(
    project_id: str,
    floor_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إضافة دور للمشروع"""
    new_floor = ProjectFloor(
        id=str(uuid.uuid4()),
        project_id=project_id,
        floor_number=floor_data.get("floor_number", 0),
        floor_name=floor_data.get("floor_name"),
        area=floor_data.get("area", 0),
        steel_factor=floor_data.get("steel_factor", 120)
    )
    
    session.add(new_floor)
    
    # تحديث المساحة الإجمالية وعدد الأدوار في المشروع
    floors_result = await session.execute(
        select(func.sum(ProjectFloor.area), func.count(ProjectFloor.id))
        .where(ProjectFloor.project_id == project_id)
    )
    total_area, floors_count = floors_result.one()
    
    await session.execute(
        update(Project).where(Project.id == project_id).values(
            total_area=(total_area or 0) + floor_data.get("area", 0),
            floors_count=(floors_count or 0) + 1
        )
    )
    
    await session.commit()
    return {"id": new_floor.id, "message": "تمت الإضافة بنجاح"}


@pg_buildings_router.put("/projects/{project_id}/floors/{floor_id}")
async def update_project_floor(
    project_id: str,
    floor_id: str,
    floor_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحديث دور في المشروع"""
    result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.id == floor_id)
    )
    floor = result.scalar_one_or_none()
    if not floor:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    
    old_area = floor.area
    
    for key in ["floor_number", "floor_name", "area", "steel_factor"]:
        if key in floor_data:
            setattr(floor, key, floor_data[key])
    
    # تحديث المساحة الإجمالية
    area_diff = floor_data.get("area", old_area) - old_area
    if area_diff != 0:
        project_result = await session.execute(
            select(Project).where(Project.id == project_id)
        )
        project = project_result.scalar_one_or_none()
        if project:
            project.total_area = (project.total_area or 0) + area_diff
    
    await session.commit()
    return {"message": "تم التحديث بنجاح"}


@pg_buildings_router.delete("/projects/{project_id}/floors/{floor_id}")
async def delete_project_floor(
    project_id: str,
    floor_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """حذف دور من المشروع"""
    # جلب المساحة قبل الحذف
    floor_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.id == floor_id)
    )
    floor = floor_result.scalar_one_or_none()
    
    if floor:
        # تحديث المساحة الإجمالية
        await session.execute(
            update(Project).where(Project.id == project_id).values(
                total_area=Project.total_area - floor.area,
                floors_count=Project.floors_count - 1
            )
        )
    
    await session.execute(delete(ProjectFloor).where(ProjectFloor.id == floor_id))
    await session.commit()
    return {"message": "تم الحذف بنجاح"}


# ==================== مواد المساحة ====================

@pg_buildings_router.get("/projects/{project_id}/area-materials")
async def get_area_materials(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """الحصول على مواد المساحة للمشروع"""
    result = await session.execute(
        select(ProjectAreaMaterial).where(ProjectAreaMaterial.project_id == project_id)
    )
    materials = result.scalars().all()
    
    return [{
        "id": m.id,
        "catalog_item_id": m.catalog_item_id,
        "item_code": m.item_code,
        "item_name": m.item_name,
        "unit": m.unit,
        "factor": m.factor,
        "unit_price": m.unit_price,
        "calculation_type": m.calculation_type,
        "selected_floors": json.loads(m.selected_floors) if m.selected_floors else [],
        "tile_width": m.tile_width,
        "tile_height": m.tile_height,
        "waste_percentage": m.waste_percentage,
        "notes": m.notes
    } for m in materials]


@pg_buildings_router.post("/projects/{project_id}/area-materials")
async def add_area_material(
    project_id: str,
    material_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إضافة مادة مساحة للمشروع"""
    # جلب بيانات المادة من الكتالوج
    catalog_item_id = material_data.get("catalog_item_id")
    catalog_result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.id == catalog_item_id)
    )
    catalog_item = catalog_result.scalar_one_or_none()
    
    new_material = ProjectAreaMaterial(
        id=str(uuid.uuid4()),
        project_id=project_id,
        catalog_item_id=catalog_item_id,
        item_code=catalog_item.item_code if catalog_item else material_data.get("item_code"),
        item_name=catalog_item.name if catalog_item else material_data.get("item_name"),
        unit=material_data.get("unit", catalog_item.unit if catalog_item else "طن"),
        factor=material_data.get("factor", 0),
        unit_price=catalog_item.price if catalog_item else material_data.get("unit_price", 0),
        calculation_type=material_data.get("calculation_type", "all_floors"),
        selected_floors=json.dumps(material_data.get("selected_floors", [])),
        tile_width=material_data.get("tile_width", 0),
        tile_height=material_data.get("tile_height", 0),
        waste_percentage=material_data.get("waste_percentage", 0),
        notes=material_data.get("notes"),
        created_at=datetime.now(timezone.utc).replace(tzinfo=None)
    )
    
    session.add(new_material)
    await session.commit()
    
    return {"id": new_material.id, "message": "تمت الإضافة بنجاح"}


@pg_buildings_router.put("/projects/{project_id}/area-materials/{material_id}")
async def update_area_material(
    project_id: str,
    material_id: str,
    material_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحديث مادة مساحة"""
    result = await session.execute(
        select(ProjectAreaMaterial).where(ProjectAreaMaterial.id == material_id)
    )
    material = result.scalar_one_or_none()
    if not material:
        raise HTTPException(status_code=404, detail="المادة غير موجودة")
    
    for key in ["factor", "unit_price", "calculation_type", "tile_width", "tile_height", "waste_percentage", "notes"]:
        if key in material_data:
            setattr(material, key, material_data[key])
    
    if "selected_floors" in material_data:
        material.selected_floors = json.dumps(material_data["selected_floors"])
    
    await session.commit()
    return {"message": "تم التحديث بنجاح"}


@pg_buildings_router.delete("/projects/{project_id}/area-materials/{material_id}")
async def delete_area_material(
    project_id: str,
    material_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """حذف مادة مساحة"""
    await session.execute(delete(ProjectAreaMaterial).where(ProjectAreaMaterial.id == material_id))
    await session.commit()
    return {"message": "تم الحذف بنجاح"}


# ==================== حساب الكميات ====================

@pg_buildings_router.get("/projects/{project_id}/calculate")
async def calculate_project_quantities(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """حساب الكميات للمشروع"""
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # جلب نماذج الوحدات
    templates_result = await session.execute(
        select(UnitTemplate).where(UnitTemplate.project_id == project_id)
    )
    templates = templates_result.scalars().all()
    
    # جلب الأدوار
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id).order_by(ProjectFloor.floor_number)
    )
    floors = floors_result.scalars().all()
    
    # جلب مواد المساحة
    area_materials_result = await session.execute(
        select(ProjectAreaMaterial).where(ProjectAreaMaterial.project_id == project_id)
    )
    area_materials = area_materials_result.scalars().all()
    
    # حساب مواد الوحدات
    materials_totals = {}
    total_units = 0
    
    for template in templates:
        # جلب مواد النموذج
        materials_result = await session.execute(
            select(UnitTemplateMaterial).where(UnitTemplateMaterial.template_id == template.id)
        )
        template_materials = materials_result.scalars().all()
        
        unit_count = template.count or 0
        total_units += unit_count
        
        for mat in template_materials:
            mat_code = mat.item_code or mat.catalog_item_id
            quantity = (mat.quantity_per_unit or 0) * unit_count
            
            if mat_code not in materials_totals:
                materials_totals[mat_code] = {
                    "catalog_item_id": mat.catalog_item_id,
                    "item_code": mat.item_code,
                    "item_name": mat.item_name,
                    "unit": mat.unit,
                    "quantity": 0,
                    "unit_price": mat.unit_price or 0,
                    "total_price": 0
                }
            
            materials_totals[mat_code]["quantity"] += quantity
    
    # حساب الأسعار
    for mat_code in materials_totals:
        mat = materials_totals[mat_code]
        mat["total_price"] = mat["quantity"] * mat["unit_price"]
    
    # حساب مواد المساحة
    total_area = project.total_area or 0
    area_materials_totals = []
    total_area_materials_cost = 0
    
    for area_mat in area_materials:
        factor = area_mat.factor or 0
        unit_price = area_mat.unit_price or 0
        calc_type = area_mat.calculation_type or "all_floors"
        selected_floors_list = json.loads(area_mat.selected_floors) if area_mat.selected_floors else []
        
        # حقول البلاط
        tile_width = area_mat.tile_width or 0
        tile_height = area_mat.tile_height or 0
        waste_percentage = area_mat.waste_percentage or 0
        is_tile = tile_width > 0 and tile_height > 0
        
        def calculate_quantity_for_area(floor_area):
            if is_tile:
                quantity = floor_area
            else:
                quantity = floor_area * factor
            
            if waste_percentage > 0:
                quantity = quantity * (1 + waste_percentage / 100)
            
            return quantity
        
        floors_detail = []
        total_quantity = 0
        
        if calc_type == "selected_floors" and selected_floors_list:
            for floor in floors:
                if floor.floor_number in selected_floors_list:
                    floor_quantity = calculate_quantity_for_area(floor.area)
                    total_quantity += floor_quantity
                    floors_detail.append({
                        "floor_number": floor.floor_number,
                        "floor_name": floor.floor_name,
                        "area": floor.area,
                        "quantity": floor_quantity
                    })
        else:
            for floor in floors:
                floor_quantity = calculate_quantity_for_area(floor.area)
                total_quantity += floor_quantity
                floors_detail.append({
                    "floor_number": floor.floor_number,
                    "floor_name": floor.floor_name,
                    "area": floor.area,
                    "quantity": floor_quantity
                })
        
        quantity = total_quantity
        total_price = quantity * unit_price
        total_area_materials_cost += total_price
        
        # حساب عدد البلاطات
        tiles_count = 0
        tiles_per_m2 = 0
        if is_tile:
            tile_area_m2 = (tile_width * tile_height) / 10000
            tiles_per_m2 = round(1 / tile_area_m2, 2) if tile_area_m2 > 0 else 0
            tiles_count = round(quantity * tiles_per_m2)
        
        area_materials_totals.append({
            "id": area_mat.id,
            "catalog_item_id": area_mat.catalog_item_id,
            "item_code": area_mat.item_code,
            "item_name": area_mat.item_name,
            "unit": "م²" if is_tile else area_mat.unit,
            "factor": factor,
            "quantity": round(quantity, 2),
            "tiles_count": tiles_count,
            "tiles_per_m2": tiles_per_m2,
            "unit_price": unit_price,
            "total_price": total_price,
            "floors_detail": floors_detail
        })
    
    # حساب الحديد
    steel_factor = project.steel_factor or 120
    total_steel_kg = total_area * steel_factor
    total_steel_tons = total_steel_kg / 1000
    
    floors_steel = []
    for floor in floors:
        floor_factor = floor.steel_factor or steel_factor
        floor_steel = floor.area * floor_factor
        floors_steel.append({
            "floor_number": floor.floor_number,
            "floor_name": floor.floor_name,
            "area": floor.area,
            "steel_factor": floor_factor,
            "steel_kg": floor_steel,
            "steel_tons": floor_steel / 1000
        })
    
    # الإجمالي
    total_unit_materials_cost = sum(m["total_price"] for m in materials_totals.values())
    grand_total = total_unit_materials_cost + total_area_materials_cost
    
    return {
        "project_id": project_id,
        "project_name": project.name,
        "total_units": total_units,
        "total_area": total_area,
        "floors_count": len(floors),
        "materials": list(materials_totals.values()),
        "area_materials": area_materials_totals,
        "steel_calculation": {
            "total_area": total_area,
            "steel_factor": steel_factor,
            "total_steel_kg": total_steel_kg,
            "total_steel_tons": round(total_steel_tons, 2),
            "floors": floors_steel
        },
        "total_unit_materials_cost": round(total_unit_materials_cost, 2),
        "total_area_materials_cost": round(total_area_materials_cost, 2),
        "total_materials_cost": round(grand_total, 2)
    }


# ==================== تتبع التوريد ====================

@pg_buildings_router.get("/projects/{project_id}/supply")
async def get_supply_tracking(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """الحصول على تتبع التوريد للمشروع"""
    result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    items = result.scalars().all()
    
    supply_list = []
    for item in items:
        required = item.required_quantity or 0
        received = item.received_quantity or 0
        remaining = required - received
        percentage = (received / required * 100) if required > 0 else 0
        
        supply_list.append({
            "id": item.id,
            "catalog_item_id": item.catalog_item_id,
            "item_code": item.item_code,
            "item_name": item.item_name,
            "unit": item.unit,
            "required_quantity": required,
            "received_quantity": received,
            "remaining_quantity": remaining,
            "completion_percentage": round(percentage, 1),
            "unit_price": item.unit_price,
            "source": item.source,
            "notes": item.notes
        })
    
    return supply_list


@pg_buildings_router.post("/projects/{project_id}/supply/sync")
async def sync_supply_tracking(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """مزامنة تتبع التوريد مع الكميات المحسوبة"""
    # حساب الكميات
    calc_result = await calculate_project_quantities(project_id, current_user, session)
    
    # حذف التوريد القديم (اختياري - يمكن تعديله للحفاظ على المستلم)
    # await session.execute(delete(SupplyTracking).where(SupplyTracking.project_id == project_id))
    
    # جلب التوريد الحالي
    existing_result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    existing_items = {item.catalog_item_id: item for item in existing_result.scalars().all()}
    
    added_count = 0
    updated_count = 0
    
    # إضافة/تحديث مواد الوحدات
    for mat in calc_result.get("materials", []):
        catalog_id = mat["catalog_item_id"]
        if catalog_id in existing_items:
            # تحديث الكمية المطلوبة فقط
            existing_items[catalog_id].required_quantity = mat["quantity"]
            existing_items[catalog_id].unit_price = mat["unit_price"]
            updated_count += 1
        else:
            # إضافة جديدة
            new_item = SupplyTracking(
                id=str(uuid.uuid4()),
                project_id=project_id,
                catalog_item_id=catalog_id,
                item_code=mat["item_code"],
                item_name=mat["item_name"],
                unit=mat["unit"],
                required_quantity=mat["quantity"],
                received_quantity=0,
                unit_price=mat["unit_price"],
                source="quantity",
                created_at=datetime.now(timezone.utc).replace(tzinfo=None)
            )
            session.add(new_item)
            added_count += 1
    
    # إضافة/تحديث مواد المساحة
    for mat in calc_result.get("area_materials", []):
        catalog_id = mat["catalog_item_id"]
        if catalog_id in existing_items:
            existing_items[catalog_id].required_quantity = mat["quantity"]
            existing_items[catalog_id].unit_price = mat["unit_price"]
            updated_count += 1
        else:
            new_item = SupplyTracking(
                id=str(uuid.uuid4()),
                project_id=project_id,
                catalog_item_id=catalog_id,
                item_code=mat["item_code"],
                item_name=mat["item_name"],
                unit=mat["unit"],
                required_quantity=mat["quantity"],
                received_quantity=0,
                unit_price=mat["unit_price"],
                source="area",
                created_at=datetime.now(timezone.utc).replace(tzinfo=None)
            )
            session.add(new_item)
            added_count += 1
    
    await session.commit()
    
    return {
        "message": "تمت المزامنة بنجاح",
        "added": added_count,
        "updated": updated_count
    }


@pg_buildings_router.put("/projects/{project_id}/supply/{item_id}")
async def update_supply_item(
    project_id: str,
    item_id: str,
    update_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحديث كمية مستلمة"""
    result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.id == item_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    
    if "received_quantity" in update_data:
        item.received_quantity = update_data["received_quantity"]
    if "notes" in update_data:
        item.notes = update_data["notes"]
    
    item.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    await session.commit()
    
    return {"message": "تم التحديث بنجاح"}


# ==================== تصدير BOQ ====================

@pg_buildings_router.get("/projects/{project_id}/export/boq-excel")
async def export_boq_excel(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير جدول الكميات BOQ إلى Excel"""
    import xlsxwriter
    
    # حساب الكميات
    calc_result = await calculate_project_quantities(project_id, current_user, session)
    
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    
    # التنسيقات
    title_format = workbook.add_format({
        'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#1e3a5f', 'font_color': 'white', 'border': 1
    })
    header_format = workbook.add_format({
        'bold': True, 'font_size': 11, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#059669', 'font_color': 'white', 'border': 1
    })
    cell_format = workbook.add_format({
        'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1
    })
    number_format = workbook.add_format({
        'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1,
        'num_format': '#,##0.00'
    })
    total_format = workbook.add_format({
        'bold': True, 'font_size': 11, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#fef3c7', 'border': 1, 'num_format': '#,##0.00'
    })
    
    # ورقة مواد الوحدات
    ws1 = workbook.add_worksheet('مواد الوحدات')
    ws1.right_to_left()
    ws1.set_column('A:A', 15)
    ws1.set_column('B:B', 35)
    ws1.set_column('C:C', 12)
    ws1.set_column('D:D', 15)
    ws1.set_column('E:E', 15)
    ws1.set_column('F:F', 18)
    
    ws1.merge_range('A1:F1', f'جدول الكميات - {project.name if project else "المشروع"}', title_format)
    ws1.merge_range('A2:F2', f'إجمالي الوحدات: {calc_result["total_units"]}', cell_format)
    
    headers = ['الكود', 'المادة', 'الوحدة', 'الكمية', 'سعر الوحدة', 'الإجمالي']
    for col, header in enumerate(headers):
        ws1.write(3, col, header, header_format)
    
    row = 4
    for mat in calc_result.get("materials", []):
        ws1.write(row, 0, mat["item_code"] or "", cell_format)
        ws1.write(row, 1, mat["item_name"], cell_format)
        ws1.write(row, 2, mat["unit"], cell_format)
        ws1.write(row, 3, mat["quantity"], number_format)
        ws1.write(row, 4, mat["unit_price"], number_format)
        ws1.write(row, 5, mat["total_price"], number_format)
        row += 1
    
    ws1.write(row, 4, 'الإجمالي', total_format)
    ws1.write(row, 5, calc_result["total_unit_materials_cost"], total_format)
    
    # ورقة مواد المساحة
    ws2 = workbook.add_worksheet('مواد المساحة')
    ws2.right_to_left()
    ws2.set_column('A:A', 15)
    ws2.set_column('B:B', 35)
    ws2.set_column('C:C', 12)
    ws2.set_column('D:D', 15)
    ws2.set_column('E:E', 15)
    ws2.set_column('F:F', 18)
    
    ws2.merge_range('A1:F1', f'مواد المساحة - {project.name if project else "المشروع"}', title_format)
    ws2.merge_range('A2:F2', f'المساحة الإجمالية: {calc_result["total_area"]} م²', cell_format)
    
    for col, header in enumerate(headers):
        ws2.write(3, col, header, header_format)
    
    row = 4
    for mat in calc_result.get("area_materials", []):
        ws2.write(row, 0, mat["item_code"] or "", cell_format)
        ws2.write(row, 1, mat["item_name"], cell_format)
        ws2.write(row, 2, mat["unit"], cell_format)
        ws2.write(row, 3, mat["quantity"], number_format)
        ws2.write(row, 4, mat["unit_price"], number_format)
        ws2.write(row, 5, mat["total_price"], number_format)
        row += 1
    
    ws2.write(row, 4, 'الإجمالي', total_format)
    ws2.write(row, 5, calc_result["total_area_materials_cost"], total_format)
    
    # ورقة الملخص
    ws3 = workbook.add_worksheet('الملخص')
    ws3.right_to_left()
    ws3.set_column('A:A', 30)
    ws3.set_column('B:B', 20)
    
    ws3.merge_range('A1:B1', 'ملخص المشروع', title_format)
    
    summary_data = [
        ('اسم المشروع', project.name if project else ""),
        ('إجمالي الوحدات', calc_result["total_units"]),
        ('المساحة الإجمالية (م²)', calc_result["total_area"]),
        ('عدد الأدوار', calc_result["floors_count"]),
        ('إجمالي الحديد (طن)', calc_result["steel_calculation"]["total_steel_tons"]),
        ('تكلفة مواد الوحدات', calc_result["total_unit_materials_cost"]),
        ('تكلفة مواد المساحة', calc_result["total_area_materials_cost"]),
        ('التكلفة الإجمالية', calc_result["total_materials_cost"]),
    ]
    
    for row, (label, value) in enumerate(summary_data, 2):
        ws3.write(row, 0, label, header_format)
        ws3.write(row, 1, value, number_format if isinstance(value, (int, float)) else cell_format)
    
    workbook.close()
    output.seek(0)
    
    filename = f"BOQ_{project_id[:8]}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== لوحة التحكم ====================

@pg_buildings_router.get("/dashboard")
async def get_buildings_dashboard(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """لوحة التحكم الرئيسية لنظام العمائر"""
    # إحصائيات عامة
    projects_count = (await session.execute(
        select(func.count()).select_from(Project)
    )).scalar() or 0
    
    templates_count = (await session.execute(
        select(func.count()).select_from(UnitTemplate)
    )).scalar() or 0
    
    total_units = (await session.execute(
        select(func.coalesce(func.sum(UnitTemplate.count), 0))
    )).scalar() or 0
    
    total_area = (await session.execute(
        select(func.coalesce(func.sum(Project.total_area), 0))
    )).scalar() or 0
    
    # ملخص المشاريع
    projects_result = await session.execute(
        select(Project).order_by(Project.created_at.desc()).limit(10)
    )
    projects = projects_result.scalars().all()
    
    projects_summary = []
    for p in projects:
        # عدد النماذج
        templates_in_project = (await session.execute(
            select(func.count()).select_from(UnitTemplate).where(UnitTemplate.project_id == p.id)
        )).scalar() or 0
        
        # عدد الوحدات
        units_in_project = (await session.execute(
            select(func.coalesce(func.sum(UnitTemplate.count), 0)).where(UnitTemplate.project_id == p.id)
        )).scalar() or 0
        
        projects_summary.append({
            "id": p.id,
            "name": p.name,
            "status": p.status,
            "templates_count": templates_in_project,
            "units_count": units_in_project,
            "area": p.total_area or 0,
            "floors_count": p.floors_count or 0
        })
    
    return {
        "total_projects": projects_count,
        "total_templates": templates_count,
        "total_units": total_units,
        "total_area": total_area,
        "projects_summary": projects_summary
    }


# ==================== تصدير واستيراد الأدوار ====================

@pg_buildings_router.get("/projects/{project_id}/export/floors-excel")
async def export_floors_excel(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير أدوار المشروع إلى Excel"""
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # جلب الأدوار
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id).order_by(ProjectFloor.floor_number)
    )
    floors = floors_result.scalars().all()
    
    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الأدوار"
    ws.sheet_view.rightToLeft = True
    
    # التنسيقات
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان
    ws.merge_cells('A1:E1')
    ws['A1'] = f'أدوار المشروع - {project.name}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align
    
    # رؤوس الأعمدة
    headers = ['رقم الدور', 'اسم الدور', 'المساحة (م²)', 'معامل التسليح (كجم/م²)', 'الحديد (طن)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # البيانات
    for row, floor in enumerate(floors, 4):
        floor_name = floor.floor_name or ""
        if floor.floor_number == -1:
            floor_name = floor_name or "اللبشة"
        elif floor.floor_number == 0:
            floor_name = floor_name or "الأرضي"
        elif floor.floor_number == 99:
            floor_name = floor_name or "السطح"
        else:
            floor_name = floor_name or f"الدور {floor.floor_number}"
        
        steel_tons = (floor.area * floor.steel_factor) / 1000
        
        ws.cell(row=row, column=1, value=floor.floor_number).border = border
        ws.cell(row=row, column=2, value=floor_name).border = border
        ws.cell(row=row, column=3, value=floor.area).border = border
        ws.cell(row=row, column=4, value=floor.steel_factor).border = border
        ws.cell(row=row, column=5, value=round(steel_tons, 2)).border = border
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 18
    
    # حفظ الملف
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Floors_{project.name}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@pg_buildings_router.post("/projects/{project_id}/import/floors")
async def import_floors_excel(
    project_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """استيراد أدوار المشروع من Excel"""
    # التحقق من نوع الملف
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف بصيغة Excel")
    
    # التحقق من وجود المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    try:
        # قراءة الملف
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        # تخطي العنوان ورؤوس الأعمدة (أول 3 صفوف)
        for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
            if not row or not row[0]:  # تخطي الصفوف الفارغة
                continue
            
            try:
                floor_number = int(row[0]) if row[0] is not None else 0
                floor_name = str(row[1]) if row[1] else ""
                area = float(row[2]) if row[2] else 0
                steel_factor = float(row[3]) if row[3] else 120
                
                # إنشاء الدور
                new_floor = ProjectFloor(
                    id=str(uuid.uuid4()),
                    project_id=project_id,
                    floor_number=floor_number,
                    floor_name=floor_name if floor_name else None,
                    area=area,
                    steel_factor=steel_factor
                )
                session.add(new_floor)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"صف {row_num}: {str(e)}")
        
        # تحديث المساحة الإجمالية وعدد الأدوار
        floors_result = await session.execute(
            select(func.sum(ProjectFloor.area), func.count(ProjectFloor.id))
            .where(ProjectFloor.project_id == project_id)
        )
        total_area, floors_count = floors_result.one()
        
        project.total_area = total_area or 0
        project.floors_count = floors_count or 0
        
        await session.commit()
        
        return {
            "message": f"تم استيراد {imported_count} دور بنجاح",
            "imported": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"فشل في قراءة الملف: {str(e)}")


# ==================== التقارير ====================

@pg_buildings_router.get("/reports/summary")
async def get_reports_summary(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير ملخص شامل لجميع المشاريع"""
    # إجماليات
    projects_count = (await session.execute(
        select(func.count()).select_from(Project)
    )).scalar() or 0
    
    templates_count = (await session.execute(
        select(func.count()).select_from(UnitTemplate)
    )).scalar() or 0
    
    total_units = (await session.execute(
        select(func.coalesce(func.sum(UnitTemplate.count), 0))
    )).scalar() or 0
    
    total_area = (await session.execute(
        select(func.coalesce(func.sum(Project.total_area), 0))
    )).scalar() or 0
    
    total_floors = (await session.execute(
        select(func.count()).select_from(ProjectFloor)
    )).scalar() or 0
    
    # تفاصيل المشاريع
    projects_result = await session.execute(
        select(Project).order_by(Project.created_at.desc())
    )
    projects = projects_result.scalars().all()
    
    projects_details = []
    for p in projects:
        # عدد الأدوار
        floors_in_project = (await session.execute(
            select(func.count()).select_from(ProjectFloor).where(ProjectFloor.project_id == p.id)
        )).scalar() or 0
        
        # إجمالي الحديد
        steel_result = await session.execute(
            select(func.sum(ProjectFloor.area * ProjectFloor.steel_factor))
            .where(ProjectFloor.project_id == p.id)
        )
        total_steel_kg = steel_result.scalar() or 0
        
        # عدد النماذج والوحدات
        templates_in_project = (await session.execute(
            select(func.count()).select_from(UnitTemplate).where(UnitTemplate.project_id == p.id)
        )).scalar() or 0
        
        units_in_project = (await session.execute(
            select(func.coalesce(func.sum(UnitTemplate.count), 0)).where(UnitTemplate.project_id == p.id)
        )).scalar() or 0
        
        projects_details.append({
            "id": p.id,
            "name": p.name,
            "code": p.code,
            "status": p.status,
            "location": p.location,
            "total_area": p.total_area or 0,
            "floors_count": floors_in_project,
            "templates_count": templates_in_project,
            "units_count": units_in_project,
            "steel_tons": round(total_steel_kg / 1000, 2) if total_steel_kg else 0,
            "created_at": p.created_at.isoformat() if p.created_at else None
        })
    
    return {
        "summary": {
            "total_projects": projects_count,
            "total_templates": templates_count,
            "total_units": total_units,
            "total_area": total_area,
            "total_floors": total_floors,
            "estimated_steel_tons": round(total_area * 120 / 1000, 2) if total_area else 0
        },
        "projects": projects_details
    }


@pg_buildings_router.get("/reports/project/{project_id}")
async def get_project_report(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير تفصيلي لمشروع واحد"""
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # جلب الأدوار
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id).order_by(ProjectFloor.floor_number)
    )
    floors = floors_result.scalars().all()
    
    # جلب النماذج
    templates_result = await session.execute(
        select(UnitTemplate).where(UnitTemplate.project_id == project_id)
    )
    templates = templates_result.scalars().all()
    
    # جلب مواد المساحة
    area_materials_result = await session.execute(
        select(ProjectAreaMaterial).where(ProjectAreaMaterial.project_id == project_id)
    )
    area_materials = area_materials_result.scalars().all()
    
    # جلب التوريد
    supply_result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    supply_items = supply_result.scalars().all()
    
    # حساب الإجماليات
    total_area = sum(f.area for f in floors)
    total_steel_kg = sum(f.area * f.steel_factor for f in floors)
    total_units = sum(t.count for t in templates)
    
    # نسبة إنجاز التوريد
    total_required = sum(s.required_quantity for s in supply_items)
    total_received = sum(s.received_quantity for s in supply_items)
    supply_completion = (total_received / total_required * 100) if total_required > 0 else 0
    
    return {
        "project": {
            "id": project.id,
            "name": project.name,
            "code": project.code,
            "location": project.location,
            "status": project.status
        },
        "summary": {
            "total_area": total_area,
            "total_floors": len(floors),
            "total_templates": len(templates),
            "total_units": total_units,
            "total_steel_kg": total_steel_kg,
            "total_steel_tons": round(total_steel_kg / 1000, 2),
            "supply_completion_percentage": round(supply_completion, 1)
        },
        "floors": [{
            "floor_number": f.floor_number,
            "floor_name": f.floor_name,
            "area": f.area,
            "steel_factor": f.steel_factor,
            "steel_tons": round(f.area * f.steel_factor / 1000, 2)
        } for f in floors],
        "templates": [{
            "code": t.code,
            "name": t.name,
            "area": t.area,
            "rooms_count": t.rooms_count,
            "count": t.count
        } for t in templates],
        "supply_status": [{
            "item_name": s.item_name,
            "required": s.required_quantity,
            "received": s.received_quantity,
            "remaining": s.required_quantity - s.received_quantity,
            "completion": round((s.received_quantity / s.required_quantity * 100) if s.required_quantity > 0 else 0, 1)
        } for s in supply_items]
    }

