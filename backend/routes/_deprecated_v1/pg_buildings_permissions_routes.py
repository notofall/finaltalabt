"""
نظام صلاحيات العمائر ودمج التوريد
Buildings Permissions & Supply Integration Routes
"""
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, update, and_, or_
from typing import Optional, List
from datetime import datetime, timezone
import uuid
import json
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database.connection import get_postgres_session
from database.models import (
    User, Project, PriceCatalogItem, PurchaseOrder, PurchaseOrderItem,
    DeliveryRecord, SupplyTracking, BuildingsPermission, UnitTemplate
)
from routes.pg_auth_routes import get_current_user_pg

pg_buildings_permissions_router = APIRouter(prefix="/api/pg/buildings")


# ==================== صلاحيات المستخدمين ====================

@pg_buildings_permissions_router.get("/permissions")
async def get_all_permissions(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """الحصول على جميع الصلاحيات (لمدير النظام أو مدير المشتريات)"""
    if current_user.role not in ["system_admin", "procurement_manager", "quantity_engineer"]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بعرض الصلاحيات")
    
    result = await session.execute(
        select(BuildingsPermission).where(BuildingsPermission.is_active == True)
    )
    permissions = result.scalars().all()
    
    return [{
        "id": p.id,
        "user_id": p.user_id,
        "user_name": p.user_name,
        "user_email": p.user_email,
        "project_id": p.project_id,
        "project_name": p.project_name,
        "can_view": p.can_view,
        "can_edit": p.can_edit,
        "can_delete": p.can_delete,
        "can_export": p.can_export,
        "granted_by_name": p.granted_by_name,
        "granted_at": p.granted_at.isoformat() if p.granted_at else None
    } for p in permissions]


@pg_buildings_permissions_router.get("/permissions/my")
async def get_my_permissions(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """الحصول على صلاحياتي في نظام العمائر"""
    result = await session.execute(
        select(BuildingsPermission).where(
            BuildingsPermission.user_id == current_user.id,
            BuildingsPermission.is_active == True
        )
    )
    permissions = result.scalars().all()
    
    # إذا كان مهندس كميات أو مدير مشتريات، له صلاحية كاملة
    if current_user.role in ["quantity_engineer", "procurement_manager", "system_admin"]:
        return {
            "has_access": True,
            "is_owner": True,
            "permissions": [{
                "project_id": None,
                "project_name": "جميع المشاريع",
                "can_view": True,
                "can_edit": True,
                "can_delete": True,
                "can_export": True
            }]
        }
    
    if not permissions:
        return {
            "has_access": False,
            "is_owner": False,
            "permissions": []
        }
    
    return {
        "has_access": True,
        "is_owner": False,
        "permissions": [{
            "id": p.id,
            "project_id": p.project_id,
            "project_name": p.project_name or "جميع المشاريع",
            "can_view": p.can_view,
            "can_edit": p.can_edit,
            "can_delete": p.can_delete,
            "can_export": p.can_export
        } for p in permissions]
    }


@pg_buildings_permissions_router.post("/permissions")
async def grant_permission(
    permission_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إعطاء صلاحية لمستخدم"""
    if current_user.role not in ["system_admin", "procurement_manager", "quantity_engineer"]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإعطاء صلاحيات")
    
    user_id = permission_data.get("user_id")
    if not user_id:
        raise HTTPException(status_code=400, detail="يجب تحديد المستخدم")
    
    # جلب المستخدم
    user_result = await session.execute(
        select(User).where(User.id == user_id)
    )
    target_user = user_result.scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # جلب المشروع إذا محدد
    project_id = permission_data.get("project_id")
    project_name = None
    if project_id:
        project_result = await session.execute(
            select(Project).where(Project.id == project_id)
        )
        project = project_result.scalar_one_or_none()
        if project:
            project_name = project.name
    
    # التحقق من عدم وجود صلاحية مكررة
    existing = await session.execute(
        select(BuildingsPermission).where(
            BuildingsPermission.user_id == user_id,
            BuildingsPermission.project_id == project_id,
            BuildingsPermission.is_active == True
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="الصلاحية موجودة مسبقاً")
    
    new_permission = BuildingsPermission(
        id=str(uuid.uuid4()),
        user_id=user_id,
        user_name=target_user.name,
        user_email=target_user.email,
        project_id=project_id,
        project_name=project_name,
        can_view=permission_data.get("can_view", True),
        can_edit=permission_data.get("can_edit", False),
        can_delete=permission_data.get("can_delete", False),
        can_export=permission_data.get("can_export", True),
        granted_by=current_user.id,
        granted_by_name=current_user.name,
        granted_at=datetime.now(timezone.utc).replace(tzinfo=None),
        is_active=True
    )
    
    session.add(new_permission)
    await session.commit()
    
    return {
        "id": new_permission.id,
        "message": f"تم إعطاء الصلاحية لـ {target_user.name} بنجاح"
    }


@pg_buildings_permissions_router.put("/permissions/{permission_id}")
async def update_permission(
    permission_id: str,
    permission_data: dict,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحديث صلاحية"""
    if current_user.role not in ["system_admin", "procurement_manager", "quantity_engineer"]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الصلاحيات")
    
    result = await session.execute(
        select(BuildingsPermission).where(BuildingsPermission.id == permission_id)
    )
    permission = result.scalar_one_or_none()
    if not permission:
        raise HTTPException(status_code=404, detail="الصلاحية غير موجودة")
    
    for key in ["can_view", "can_edit", "can_delete", "can_export"]:
        if key in permission_data:
            setattr(permission, key, permission_data[key])
    
    await session.commit()
    return {"message": "تم تحديث الصلاحية بنجاح"}


@pg_buildings_permissions_router.delete("/permissions/{permission_id}")
async def revoke_permission(
    permission_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إلغاء صلاحية"""
    if current_user.role not in ["system_admin", "procurement_manager", "quantity_engineer"]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإلغاء الصلاحيات")
    
    result = await session.execute(
        select(BuildingsPermission).where(BuildingsPermission.id == permission_id)
    )
    permission = result.scalar_one_or_none()
    if not permission:
        raise HTTPException(status_code=404, detail="الصلاحية غير موجودة")
    
    permission.is_active = False
    await session.commit()
    
    return {"message": "تم إلغاء الصلاحية بنجاح"}


@pg_buildings_permissions_router.get("/users/available")
async def get_available_users(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """الحصول على المستخدمين المتاحين لإعطائهم صلاحيات"""
    result = await session.execute(
        select(User).where(User.is_active == True)
    )
    users = result.scalars().all()
    
    return [{
        "id": u.id,
        "name": u.name,
        "email": u.email,
        "role": u.role
    } for u in users]


# ==================== ربط التوريد مع استلام الطلبات ====================

@pg_buildings_permissions_router.post("/supply/sync-from-delivery")
async def sync_supply_from_deliveries(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """مزامنة التوريد من سجلات التسليم - يخصم الكميات المستلمة تلقائياً"""
    # جلب جميع أوامر الشراء المستلمة للمشروع
    orders_result = await session.execute(
        select(PurchaseOrder).where(
            PurchaseOrder.project_id == project_id,
            PurchaseOrder.status == "delivered"
        )
    )
    delivered_orders = orders_result.scalars().all()
    
    updated_count = 0
    
    for order in delivered_orders:
        # جلب عناصر الطلب
        items_result = await session.execute(
            select(PurchaseOrderItem).where(PurchaseOrderItem.purchase_order_id == order.id)
        )
        order_items = items_result.scalars().all()
        
        for item in order_items:
            if not item.catalog_item_id:
                continue
            
            # البحث عن العنصر في تتبع التوريد
            supply_result = await session.execute(
                select(SupplyTracking).where(
                    SupplyTracking.project_id == project_id,
                    SupplyTracking.catalog_item_id == item.catalog_item_id
                )
            )
            supply_item = supply_result.scalar_one_or_none()
            
            if supply_item:
                # تحديث الكمية المستلمة
                supply_item.received_quantity = (supply_item.received_quantity or 0) + item.quantity
                supply_item.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
                updated_count += 1
    
    await session.commit()
    
    return {
        "message": f"تم تحديث {updated_count} عنصر من التوريد",
        "updated_count": updated_count
    }


# ==================== استيراد وتصدير المشاريع ====================

@pg_buildings_permissions_router.get("/export/project-template")
async def download_project_template(
    current_user: User = Depends(get_current_user_pg)
):
    """تحميل نموذج استيراد المشروع"""
    wb = openpyxl.Workbook()
    
    # التنسيقات
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
    note_fill = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
    
    # ورقة الأدوار
    ws_floors = wb.active
    ws_floors.title = "الأدوار"
    ws_floors.sheet_view.rightToLeft = True
    
    floors_headers = ['رقم الدور', 'اسم الدور', 'المساحة (م²)', 'معامل التسليح (كجم/م²)']
    for col, header in enumerate(floors_headers, 1):
        cell = ws_floors.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # أمثلة
    example_floors = [
        [-1, 'اللبشة', 500, 150],
        [0, 'الأرضي', 450, 120],
        [1, 'الدور الأول', 400, 100],
        [2, 'الدور الثاني', 400, 100],
    ]
    for row, floor in enumerate(example_floors, 2):
        for col, val in enumerate(floor, 1):
            cell = ws_floors.cell(row=row, column=col, value=val)
            cell.fill = example_fill
    
    # ملاحظات
    ws_floors.cell(row=7, column=1, value="ملاحظات:").fill = note_fill
    ws_floors.cell(row=8, column=1, value="• رقم الدور: -1 للبشة، 0 للأرضي، 1-15 للأدوار، 99 للسطح")
    ws_floors.cell(row=9, column=1, value="• معامل التسليح: عادة 100-150 كجم/م²")
    
    ws_floors.column_dimensions['A'].width = 15
    ws_floors.column_dimensions['B'].width = 20
    ws_floors.column_dimensions['C'].width = 18
    ws_floors.column_dimensions['D'].width = 25
    
    # ورقة نماذج الوحدات
    ws_templates = wb.create_sheet("نماذج الوحدات")
    ws_templates.sheet_view.rightToLeft = True
    
    templates_headers = ['كود النموذج', 'اسم النموذج', 'المساحة (م²)', 'عدد الغرف', 'عدد الحمامات', 'عدد الوحدات']
    for col, header in enumerate(templates_headers, 1):
        cell = ws_templates.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # أمثلة
    example_templates = [
        ['UNIT-A', 'شقة 3 غرف', 120, 3, 2, 8],
        ['UNIT-B', 'شقة 4 غرف', 150, 4, 3, 4],
        ['UNIT-C', 'استوديو', 60, 1, 1, 6],
    ]
    for row, template in enumerate(example_templates, 2):
        for col, val in enumerate(template, 1):
            cell = ws_templates.cell(row=row, column=col, value=val)
            cell.fill = example_fill
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_templates.column_dimensions[col].width = 18
    
    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=project_import_template.xlsx"}
    )


@pg_buildings_permissions_router.post("/import/project/{project_id}")
async def import_project_data(
    project_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """استيراد بيانات المشروع من Excel (الأدوار ونماذج الوحدات)"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف بصيغة Excel")
    
    # التحقق من المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        results = {
            "floors_imported": 0,
            "templates_imported": 0,
            "errors": []
        }
        
        # استيراد الأدوار
        if "الأدوار" in wb.sheetnames:
            ws_floors = wb["الأدوار"]
            from database.models import ProjectFloor
            
            for row_num, row in enumerate(ws_floors.iter_rows(min_row=2, values_only=True), 2):
                if not row or row[0] is None:
                    continue
                
                try:
                    floor_number = int(row[0])
                    floor_name = str(row[1]) if row[1] else None
                    area = float(row[2]) if row[2] else 0
                    steel_factor = float(row[3]) if row[3] else 120
                    
                    new_floor = ProjectFloor(
                        id=str(uuid.uuid4()),
                        project_id=project_id,
                        floor_number=floor_number,
                        floor_name=floor_name,
                        area=area,
                        steel_factor=steel_factor
                    )
                    session.add(new_floor)
                    results["floors_imported"] += 1
                except Exception as e:
                    results["errors"].append(f"الأدوار - صف {row_num}: {str(e)}")
        
        # استيراد نماذج الوحدات
        if "نماذج الوحدات" in wb.sheetnames:
            ws_templates = wb["نماذج الوحدات"]
            
            for row_num, row in enumerate(ws_templates.iter_rows(min_row=2, values_only=True), 2):
                if not row or not row[0]:
                    continue
                
                try:
                    new_template = UnitTemplate(
                        id=str(uuid.uuid4()),
                        code=str(row[0]),
                        name=str(row[1]) if row[1] else str(row[0]),
                        area=float(row[2]) if row[2] else 0,
                        rooms_count=int(row[3]) if row[3] else 0,
                        bathrooms_count=int(row[4]) if row[4] else 0,
                        count=int(row[5]) if row[5] else 0,
                        project_id=project_id,
                        project_name=project.name,
                        created_by=current_user.id,
                        created_by_name=current_user.name,
                        created_at=datetime.now(timezone.utc).replace(tzinfo=None)
                    )
                    session.add(new_template)
                    results["templates_imported"] += 1
                except Exception as e:
                    results["errors"].append(f"النماذج - صف {row_num}: {str(e)}")
        
        # تحديث إحصائيات المشروع
        from database.models import ProjectFloor
        floors_result = await session.execute(
            select(func.sum(ProjectFloor.area), func.count(ProjectFloor.id))
            .where(ProjectFloor.project_id == project_id)
        )
        total_area, floors_count = floors_result.one()
        project.total_area = total_area or 0
        project.floors_count = floors_count or 0
        
        await session.commit()
        
        return {
            "message": f"تم استيراد {results['floors_imported']} دور و {results['templates_imported']} نموذج",
            **results
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"فشل في قراءة الملف: {str(e)}")


# ==================== تقارير متقدمة ====================

@pg_buildings_permissions_router.get("/reports/supply-details/{project_id}")
async def get_supply_details_report(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير تفصيلي للتوريد"""
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # جلب بيانات التوريد
    supply_result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    supply_items = supply_result.scalars().all()
    
    # حساب الإجماليات
    total_required = sum(s.required_quantity for s in supply_items)
    total_received = sum(s.received_quantity for s in supply_items)
    total_remaining = total_required - total_received
    overall_completion = (total_received / total_required * 100) if total_required > 0 else 0
    
    # تصنيف حسب حالة الإنجاز
    completed_items = []
    in_progress_items = []
    not_started_items = []
    
    for item in supply_items:
        required = item.required_quantity or 0
        received = item.received_quantity or 0
        remaining = required - received
        completion = (received / required * 100) if required > 0 else 0
        
        item_data = {
            "id": item.id,
            "item_code": item.item_code,
            "item_name": item.item_name,
            "unit": item.unit,
            "required_quantity": required,
            "received_quantity": received,
            "remaining_quantity": remaining,
            "completion_percentage": round(completion, 1),
            "unit_price": item.unit_price or 0,
            "required_value": required * (item.unit_price or 0),
            "received_value": received * (item.unit_price or 0),
            "remaining_value": remaining * (item.unit_price or 0),
            "source": item.source
        }
        
        if completion >= 100:
            completed_items.append(item_data)
        elif completion > 0:
            in_progress_items.append(item_data)
        else:
            not_started_items.append(item_data)
    
    # ترتيب حسب نسبة الإنجاز
    in_progress_items.sort(key=lambda x: x["completion_percentage"], reverse=True)
    
    return {
        "project": {
            "id": project.id,
            "name": project.name
        },
        "summary": {
            "total_items": len(supply_items),
            "completed_count": len(completed_items),
            "in_progress_count": len(in_progress_items),
            "not_started_count": len(not_started_items),
            "total_required": total_required,
            "total_received": total_received,
            "total_remaining": total_remaining,
            "overall_completion": round(overall_completion, 1),
            "total_required_value": sum(s.required_quantity * (s.unit_price or 0) for s in supply_items),
            "total_received_value": sum(s.received_quantity * (s.unit_price or 0) for s in supply_items),
        },
        "completed_items": completed_items,
        "in_progress_items": in_progress_items,
        "not_started_items": not_started_items
    }


@pg_buildings_permissions_router.get("/reports/supply-export/{project_id}")
async def export_supply_report(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير تقرير التوريد إلى Excel"""
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # جلب بيانات التوريد
    supply_result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    supply_items = supply_result.scalars().all()
    
    # إنشاء Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير التوريد"
    ws.sheet_view.rightToLeft = True
    
    # التنسيقات
    title_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    completed_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    progress_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    not_started_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    
    # العنوان
    ws.merge_cells('A1:I1')
    ws['A1'] = f'تقرير التوريد - {project.name}'
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # رؤوس الأعمدة
    headers = ['الكود', 'المادة', 'الوحدة', 'المطلوب', 'المستلم', 'المتبقي', 'الإنجاز %', 'السعر', 'القيمة المتبقية']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # البيانات
    row = 4
    for item in supply_items:
        required = item.required_quantity or 0
        received = item.received_quantity or 0
        remaining = required - received
        completion = (received / required * 100) if required > 0 else 0
        remaining_value = remaining * (item.unit_price or 0)
        
        # تحديد لون الصف
        if completion >= 100:
            row_fill = completed_fill
        elif completion > 0:
            row_fill = progress_fill
        else:
            row_fill = not_started_fill
        
        data = [
            item.item_code or "",
            item.item_name,
            item.unit,
            required,
            received,
            remaining,
            f"{completion:.1f}%",
            item.unit_price or 0,
            remaining_value
        ]
        
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
        
        row += 1
    
    # الإجماليات
    total_row = row
    ws.cell(row=total_row, column=2, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(s.required_quantity or 0 for s in supply_items)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(s.received_quantity or 0 for s in supply_items)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum((s.required_quantity or 0) - (s.received_quantity or 0) for s in supply_items)).font = Font(bold=True)
    
    total_required = sum(s.required_quantity or 0 for s in supply_items)
    total_received = sum(s.received_quantity or 0 for s in supply_items)
    overall_completion = (total_received / total_required * 100) if total_required > 0 else 0
    ws.cell(row=total_row, column=7, value=f"{overall_completion:.1f}%").font = Font(bold=True)
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    
    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Supply_Report_{project.name}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )
