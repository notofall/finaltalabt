"""
Catalog API v2 - Using Service Layer
V2 كتالوج API - باستخدام طبقة الخدمات

Architecture: Route -> Service -> Repository
"""
from fastapi import APIRouter, HTTPException, Depends, status, Query
from pydantic import BaseModel
from typing import Optional, List

from app.services import CatalogService
from app.dependencies import get_catalog_service
from app.config import PaginationConfig
from routes.v2_auth_routes import get_current_user, require_admin

# Create router
router = APIRouter(prefix="/api/v2/catalog", tags=["V2 Catalog"])

DEFAULT_LIMIT = PaginationConfig.DEFAULT_PAGE_SIZE
MAX_LIMIT = PaginationConfig.MAX_PAGE_SIZE


# ==================== PYDANTIC MODELS ====================

class CatalogItemCreate(BaseModel):
    name: str
    unit: str
    price: float
    category_name: Optional[str] = None
    category_code: Optional[str] = None  # كود التصنيف لتوليد كود الصنف
    item_code: Optional[str] = None
    description: Optional[str] = None
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None


class CatalogItemUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    category_name: Optional[str] = None
    description: Optional[str] = None


class CatalogItemResponse(BaseModel):
    id: str
    item_code: Optional[str] = None
    name: str
    unit: str
    price: float
    category_name: Optional[str] = None
    description: Optional[str] = None
    supplier_name: Optional[str] = None
    is_active: bool = True
    created_at: Optional[str] = None
    
    class Config:
        from_attributes = True


class CatalogListResponse(BaseModel):
    """Paginated catalog response"""
    items: List[CatalogItemResponse]
    total: int
    skip: int
    limit: int
    has_more: bool


class AliasCreate(BaseModel):
    alias_name: str
    catalog_item_id: str


class AliasResponse(BaseModel):
    id: str
    alias_name: str
    catalog_item_id: str
    catalog_item_name: Optional[str] = None
    usage_count: int = 0
    created_at: Optional[str] = None
    
    class Config:
        from_attributes = True


# ==================== HELPER ====================

def item_to_response(item) -> CatalogItemResponse:
    """Convert PriceCatalogItem to response"""
    return CatalogItemResponse(
        id=str(item.id),
        item_code=item.item_code,
        name=item.name,
        unit=item.unit,
        price=item.price,
        category_name=item.category_name,
        description=item.description,
        supplier_name=item.supplier_name,
        is_active=item.is_active,
        created_at=item.created_at.isoformat() if item.created_at else None
    )


def alias_to_response(alias) -> AliasResponse:
    """Convert ItemAlias to response"""
    return AliasResponse(
        id=str(alias.id),
        alias_name=alias.alias_name,
        catalog_item_id=str(alias.catalog_item_id),
        catalog_item_name=getattr(alias, 'catalog_item_name', None),
        usage_count=getattr(alias, 'usage_count', 0),
        created_at=alias.created_at.isoformat() if alias.created_at else None
    )


# ==================== CATALOG ITEMS ====================

@router.get("/items", response_model=CatalogListResponse)
async def get_catalog_items(
    skip: int = Query(0, ge=0),
    limit: int = Query(DEFAULT_LIMIT, ge=1, le=MAX_LIMIT),
    category: Optional[str] = None,
    search: Optional[str] = None,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Get catalog items with pagination and filtering
    Uses: CatalogService -> CatalogRepository
    """
    limit = min(limit, MAX_LIMIT)
    total = await catalog_service.count_items()
    
    if search:
        items = await catalog_service.search_items(search, limit)
    elif category:
        items = await catalog_service.get_items_by_category(category)
        items = items[skip:skip + limit]
    else:
        items = await catalog_service.get_all_items(skip, limit)
    
    return CatalogListResponse(
        items=[item_to_response(i) for i in items],
        total=total,
        skip=skip,
        limit=limit,
        has_more=(skip + len(items)) < total
    )


@router.get("/items/{item_id}", response_model=CatalogItemResponse)
async def get_catalog_item(
    item_id: str,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Get catalog item by ID
    Uses: CatalogService -> CatalogRepository
    """
    item = await catalog_service.get_item_by_id(item_id)
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الصنف غير موجود"
        )
    return item_to_response(item)


@router.get("/items/code/{code}", response_model=CatalogItemResponse)
async def get_catalog_item_by_code(
    code: str,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Get catalog item by code
    Uses: CatalogService -> CatalogRepository
    """
    item = await catalog_service.get_item_by_code(code)
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الصنف غير موجود"
        )
    return item_to_response(item)


@router.post("/items", response_model=CatalogItemResponse, status_code=status.HTTP_201_CREATED)
async def create_catalog_item(
    data: CatalogItemCreate,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Create catalog item
    Uses: CatalogService -> CatalogRepository
    """
    try:
        item = await catalog_service.create_item(
            name=data.name,
            unit=data.unit,
            price=data.price,
            category_name=data.category_name,
            category_code=data.category_code,
            item_code=data.item_code,
            description=data.description,
            supplier_id=data.supplier_id,
            supplier_name=data.supplier_name,
            created_by=str(current_user.id),
            created_by_name=current_user.name
        )
        return item_to_response(item)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/items/{item_id}", response_model=CatalogItemResponse)
async def update_catalog_item(
    item_id: str,
    data: CatalogItemUpdate,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Update catalog item
    Uses: CatalogService -> CatalogRepository
    """
    update_data = data.model_dump(exclude_unset=True)
    item = await catalog_service.update_item(item_id, update_data)
    
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الصنف غير موجود"
        )
    return item_to_response(item)


@router.delete("/items/{item_id}")
async def delete_catalog_item(
    item_id: str,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Delete catalog item
    Uses: CatalogService -> CatalogRepository
    """
    success = await catalog_service.delete_item(item_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الصنف غير موجود"
        )
    return {"message": "تم حذف الصنف بنجاح"}


@router.get("/categories", response_model=List[str])
async def get_categories(
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Get distinct categories
    Uses: CatalogService -> CatalogRepository
    """
    return await catalog_service.get_categories()


@router.get("/suggest-code")
async def suggest_item_code(
    category: Optional[str] = Query(None, description="Category name"),
    category_code: Optional[str] = Query(None, description="Category code for generating item code"),
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Suggest next item code based on category code or category name
    If category_code provided (e.g., "1"), returns codes like "1-0001", "1-0002"
    Returns suggested code that can be modified by user
    """
    if category_code:
        suggested_code = await catalog_service.catalog_repo.get_next_code_by_category(category_code, category)
    else:
        suggested_code = await catalog_service.catalog_repo.get_next_code(category)
    return {"suggested_code": suggested_code, "category": category, "category_code": category_code}


@router.get("/search", response_model=List[CatalogItemResponse])
async def search_catalog(
    q: str = Query(..., min_length=1),
    limit: int = Query(50, ge=1, le=100),
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Search catalog items
    Uses: CatalogService -> CatalogRepository
    """
    items = await catalog_service.search_items(q, limit)
    return [item_to_response(i) for i in items]


# ==================== ITEM ALIASES ====================

@router.get("/aliases", response_model=List[AliasResponse])
async def get_aliases(
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Get all item aliases
    Uses: CatalogService -> CatalogRepository
    """
    aliases = await catalog_service.get_all_aliases()
    return [alias_to_response(a) for a in aliases]


@router.post("/aliases", response_model=AliasResponse, status_code=status.HTTP_201_CREATED)
async def create_alias(
    data: AliasCreate,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Create item alias
    Uses: CatalogService -> CatalogRepository
    """
    # Check if item exists
    item = await catalog_service.get_item_by_id(data.catalog_item_id)
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الصنف المرتبط غير موجود"
        )
    
    alias = await catalog_service.create_alias(
        alias_name=data.alias_name,
        catalog_item_id=data.catalog_item_id,
        created_by=str(current_user.id)
    )
    return alias_to_response(alias)


@router.delete("/aliases/{alias_id}")
async def delete_alias(
    alias_id: str,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Delete item alias
    Uses: CatalogService -> CatalogRepository
    """
    success = await catalog_service.delete_alias(alias_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الاسم البديل غير موجود"
        )
    return {"message": "تم حذف الاسم البديل بنجاح"}


@router.get("/suggest/{item_name}")
async def suggest_standard_name(
    item_name: str,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """
    Suggest standard name for an item
    Uses: CatalogService -> CatalogRepository
    """
    return await catalog_service.suggest_standard_name(item_name)


# ==================== IMPORT/EXPORT ENDPOINTS ====================

from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import io
import csv

@router.get("/template")
async def get_import_template(
    current_user = Depends(get_current_user)
):
    """Get CSV template for importing catalog items"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["name", "unit", "price", "category_name", "item_code", "description", "supplier_name"])
    
    # Example rows
    writer.writerow(["اسمنت بورتلاند", "طن", "450", "مواد البناء", "CEM-001", "اسمنت عادي", "شركة الأسمنت"])
    writer.writerow(["حديد تسليح 12مم", "طن", "3500", "حديد", "STEEL-12", "حديد تسليح قطر 12 مم", "حديد السعودية"])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=catalog_template.csv"}
    )


@router.post("/import")
async def import_catalog(
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """Import catalog items from CSV file"""
    require_admin(current_user)
    
    content = await file.read()
    
    try:
        decoded = content.decode('utf-8-sig')
    except:
        decoded = content.decode('utf-8')
    
    reader = csv.DictReader(io.StringIO(decoded))
    
    imported = 0
    errors = []
    
    for row_num, row in enumerate(reader, start=2):
        try:
            name = row.get('name', '').strip()
            if not name:
                continue
            
            await catalog_service.create_item(
                name=name,
                unit=row.get('unit', 'قطعة').strip(),
                price=float(row.get('price', 0)),
                category_name=row.get('category_name', '').strip() or None,
                item_code=row.get('item_code', '').strip() or None,
                description=row.get('description', '').strip() or None,
                supplier_id=None,
                supplier_name=row.get('supplier_name', '').strip() or None
            )
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    return {
        "message": f"تم استيراد {imported} صنف",
        "imported": imported,
        "errors": errors[:10]  # Return first 10 errors
    }


@router.get("/export")
async def export_catalog(
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """Export catalog items to CSV"""
    items_orm = await catalog_service.get_all_items(skip=0, limit=10000)
    # Convert ORM to dict
    items = [item_to_response(i).model_dump() for i in items_orm]
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["كود الصنف", "اسم الصنف", "الوحدة", "السعر", "الفئة", "المورد", "الوصف"])
    
    for item in items:
        writer.writerow([
            item.get("item_code", ""),
            item.get("name", ""),
            item.get("unit", ""),
            item.get("price", 0),
            item.get("category_name", ""),
            item.get("supplier_name", ""),
            item.get("description", "")
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=price_catalog.csv"}
    )


@router.get("/export/excel")
async def export_catalog_excel(
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """Export catalog items to Excel format (XLSX)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from urllib.parse import quote
        
        items_orm = await catalog_service.get_all_items(skip=0, limit=10000)
        items = [item_to_response(i).model_dump() for i in items_orm]
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "كتالوج الأسعار"
        ws.sheet_view.rightToLeft = True
        
        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["كود الصنف", "اسم الصنف", "الوحدة", "السعر", "الفئة", "المورد", "الوصف"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=item.get("item_code", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=item.get("name", "")).border = thin_border
            ws.cell(row=row_idx, column=3, value=item.get("unit", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=item.get("price", 0)).border = thin_border
            ws.cell(row=row_idx, column=5, value=item.get("category_name", "") or "").border = thin_border
            ws.cell(row=row_idx, column=6, value=item.get("supplier_name", "") or "").border = thin_border
            ws.cell(row=row_idx, column=7, value=item.get("description", "") or "").border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 30
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"كتالوج_الأسعار_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.xlsx"
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")


@router.post("/validate-items")
async def validate_items(
    data: dict,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """Validate items against catalog - يتحقق من الربط أو الاسم البديل"""
    items = data.get("items", [])
    
    results = []
    for item in items:
        item_name = item.get("name", "")
        catalog_item_id = item.get("catalog_item_id")  # الربط المباشر بالكتالوج
        
        # إذا كان الصنف مربوط بالكتالوج مسبقاً
        if catalog_item_id:
            catalog_item_orm = await catalog_service.get_item_by_id(catalog_item_id)
            if catalog_item_orm:
                catalog_item = item_to_response(catalog_item_orm).model_dump()
                results.append({
                    "name": item_name,
                    "found": True,
                    "linked": True,  # مربوط بالكتالوج
                    "catalog_item": catalog_item,
                    "price_match": True
                })
                continue
        
        # البحث عن الصنف بالاسم أو الاسم البديل
        search_result_orm = await catalog_service.search_items(item_name, limit=1)
        
        # البحث في الأسماء البديلة أيضاً
        if not search_result_orm:
            alias = await catalog_service.find_alias(item_name)
            if alias:
                linked_item = await catalog_service.get_item_by_id(alias.catalog_item_id)
                if linked_item:
                    search_result_orm = [linked_item]
        
        if search_result_orm and search_result_orm[0]:
            catalog_item = item_to_response(search_result_orm[0]).model_dump()
            results.append({
                "name": item_name,
                "found": True,
                "linked": False,
                "catalog_item": catalog_item,
                "price_match": abs(catalog_item.get("price", 0) - item.get("unit_price", 0)) < 0.01
            })
        else:
            results.append({
                "name": item_name,
                "found": False,
                "linked": False,
                "catalog_item": None,
                "price_match": False
            })
    
    all_valid = all(r["found"] or r.get("linked", False) for r in results)
    
    return {
        "items": results,
        "all_valid": all_valid,
        "total_found": sum(1 for r in results if r["found"]),
        "total_not_found": sum(1 for r in results if not r["found"])
    }


@router.post("/check-best-price")
async def check_best_price(
    item_name: str = Query(...),
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """Check best price for an item from catalog"""
    results_orm = await catalog_service.search_items(item_name, limit=5)
    
    if not results_orm:
        return {"found": False, "message": "لم يتم العثور على صنف مطابق"}
    
    # Convert ORM to dict
    results = [item_to_response(i).model_dump() for i in results_orm]
    
    # Find best (lowest) price
    best = min(results, key=lambda x: x.get("price", float('inf')))
    
    return {
        "found": True,
        "best_price": best.get("price"),
        "best_supplier": best.get("supplier_name"),
        "item_name": best.get("name"),
        "alternatives": results
    }


@router.post("/quick-add")
async def quick_add_item(
    data: dict,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """Quick add item to catalog"""
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="اسم الصنف مطلوب")
    
    # Get user info
    user_id = current_user.get("id") if isinstance(current_user, dict) else str(current_user.id)
    user_name = current_user.get("name", "") if isinstance(current_user, dict) else current_user.name
    
    item_orm = await catalog_service.create_item(
        name=name,
        unit=data.get("unit", "قطعة"),
        price=float(data.get("price", 0)),
        category_name=data.get("category_name"),
        item_code=data.get("item_code"),
        description=data.get("description"),
        supplier_id=data.get("supplier_id"),
        supplier_name=data.get("supplier_name"),
        created_by=user_id,
        created_by_name=user_name
    )
    
    # Convert ORM to dict
    item = item_to_response(item_orm).model_dump()
    
    return {"message": "تم إضافة الصنف بنجاح", "item": item}


@router.get("/aliases/suggest/{item_name}")
async def suggest_alias(
    item_name: str,
    current_user = Depends(get_current_user),
    catalog_service: CatalogService = Depends(get_catalog_service)
):
    """Suggest standard name for an item (alias lookup)"""
    return await catalog_service.suggest_standard_name(item_name)
