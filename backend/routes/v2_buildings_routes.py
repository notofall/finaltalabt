"""
Buildings API v2 - Using Service Layer
V2 مباني API - باستخدام طبقة الخدمات

Architecture: Route -> Service -> Repository
"""
from fastapi import APIRouter, HTTPException, Depends, status, File, UploadFile
from pydantic import BaseModel
from typing import Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from uuid import uuid4

from app.services import BuildingsService
from app.dependencies import get_buildings_service
from routes.v2_auth_routes import get_current_user
from database.connection import get_postgres_session
from database.models import Project

# Create router
router = APIRouter(prefix="/api/v2/buildings", tags=["V2 Buildings"])


# ==================== PYDANTIC MODELS ====================

class TemplateCreate(BaseModel):
    code: str
    name: str
    area: float = 0
    rooms_count: int = 0
    bathrooms_count: int = 0
    count: int = 1
    description: Optional[str] = None


class TemplateUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    area: Optional[float] = None
    rooms_count: Optional[int] = None
    bathrooms_count: Optional[int] = None
    count: Optional[int] = None
    description: Optional[str] = None


class TemplateMaterialCreate(BaseModel):
    catalog_item_id: str
    item_code: str
    item_name: str
    unit: str
    quantity_per_unit: float
    unit_price: float = 0


class FloorCreate(BaseModel):
    floor_number: int
    floor_name: str
    area: float
    steel_factor: float = 100


class FloorUpdate(BaseModel):
    floor_name: Optional[str] = None
    area: Optional[float] = None
    steel_factor: Optional[float] = None


class AreaMaterialCreate(BaseModel):
    catalog_item_id: str
    item_name: str
    unit: str
    # طريقة الحساب: factor (بالمعامل) أو direct (كمية مباشرة)
    calculation_method: str = "factor"
    factor: float = 0  # المعامل (للطريقة factor)
    direct_quantity: float = 0  # الكمية المباشرة (للطريقة direct)
    unit_price: float = 0
    # نوع الحساب: all_floors أو selected_floor
    calculation_type: str = "all_floors"
    selected_floor_id: Optional[str] = None  # ID الدور المحدد
    # للبلاط
    tile_width: float = 0  # عرض البلاطة بالسم
    tile_height: float = 0  # طول البلاطة بالسم
    waste_percentage: float = 0  # نسبة الهالك %
    notes: Optional[str] = None


class AreaMaterialUpdate(BaseModel):
    item_name: Optional[str] = None
    calculation_method: Optional[str] = None
    factor: Optional[float] = None
    direct_quantity: Optional[float] = None
    unit_price: Optional[float] = None
    calculation_type: Optional[str] = None
    selected_floor_id: Optional[str] = None
    tile_width: Optional[float] = None
    tile_height: Optional[float] = None
    waste_percentage: Optional[float] = None
    notes: Optional[str] = None


class SupplyUpdate(BaseModel):
    received_quantity: Optional[float] = None
    notes: Optional[str] = None


# ==================== HELPER ====================

def floor_to_response(floor) -> dict:
    """Convert ProjectFloor to response"""
    return {
        "id": str(floor.id),
        "project_id": str(floor.project_id),
        "floor_number": floor.floor_number,
        "floor_name": floor.floor_name,
        "area": floor.area,
        "steel_factor": floor.steel_factor
    }


def area_material_to_response(material) -> dict:
    """Convert ProjectAreaMaterial to response"""
    return {
        "id": str(material.id),
        "project_id": str(material.project_id),
        "catalog_item_id": str(material.catalog_item_id) if material.catalog_item_id else None,
        "item_code": getattr(material, 'item_code', None),
        "item_name": material.item_name,
        "unit": material.unit,
        "calculation_method": getattr(material, 'calculation_method', 'factor') or 'factor',
        "factor": material.factor,
        "direct_quantity": getattr(material, 'direct_quantity', 0) or 0,
        "unit_price": material.unit_price,
        "calculation_type": getattr(material, 'calculation_type', 'all_floors') or 'all_floors',
        "selected_floor_id": getattr(material, 'selected_floor_id', None),
        "tile_width": getattr(material, 'tile_width', 0) or 0,
        "tile_height": getattr(material, 'tile_height', 0) or 0,
        "waste_percentage": getattr(material, 'waste_percentage', 0) or 0,
        "notes": getattr(material, 'notes', None),
        "created_at": material.created_at.isoformat() if material.created_at else None
    }


def supply_to_response(item) -> dict:
    """Convert SupplyTracking to response"""
    remaining = (item.required_quantity or 0) - (item.received_quantity or 0)
    completion = round((item.received_quantity or 0) / item.required_quantity * 100, 1) if item.required_quantity else 0
    
    return {
        "id": str(item.id),
        "project_id": str(item.project_id),
        "item_name": item.item_name,
        "unit": item.unit,
        "required_quantity": item.required_quantity,
        "received_quantity": item.received_quantity or 0,
        "remaining_quantity": remaining,
        "completion_percentage": completion,
        "notes": item.notes
    }


# ==================== DASHBOARD & REPORTS ====================

@router.get("/projects")
async def get_buildings_projects(
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get projects enabled for buildings system"""
    # Get only projects with is_building_project = True
    result = await session.execute(
        select(Project).where(
            Project.status == "active"
        ).order_by(Project.created_at.desc())
    )
    projects = result.scalars().all()
    
    # Filter projects that have is_building_project = True (default is True for new ones)
    building_projects = []
    for p in projects:
        # Check if is_building_project exists and is True (or default to True if field is new)
        is_building = getattr(p, 'is_building_project', True)
        if is_building is None or is_building:
            building_projects.append({
                "id": str(p.id),
                "name": p.name,
                "code": p.code,
                "owner_name": p.owner_name,
                "status": p.status,
                "total_area": p.total_area or 0,
                "floors_count": p.floors_count or 0
            })
    
    return building_projects


@router.delete("/projects/{project_id}")
async def delete_project_quantities(
    project_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Delete all quantities data for a project (BOQ, templates, floors, supply tracking)"""
    from database.models import UnitTemplate, UnitTemplateMaterial, ProjectFloor, ProjectAreaMaterial, SupplyTracking
    
    result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Get template IDs for this project to delete their materials
    templates_result = await session.execute(
        select(UnitTemplate.id).where(UnitTemplate.project_id == project_id)
    )
    template_ids = [t[0] for t in templates_result.fetchall()]
    
    deleted_counts = {
        "template_materials": 0,
        "templates": 0,
        "floors": 0,
        "area_materials": 0,
        "supply_tracking": 0
    }
    
    # Delete template materials first (foreign key constraint)
    if template_ids:
        del_materials = await session.execute(
            delete(UnitTemplateMaterial).where(UnitTemplateMaterial.template_id.in_(template_ids))
        )
        deleted_counts["template_materials"] = del_materials.rowcount
    
    # Delete unit templates
    del_templates = await session.execute(
        delete(UnitTemplate).where(UnitTemplate.project_id == project_id)
    )
    deleted_counts["templates"] = del_templates.rowcount
    
    # Delete project floors
    del_floors = await session.execute(
        delete(ProjectFloor).where(ProjectFloor.project_id == project_id)
    )
    deleted_counts["floors"] = del_floors.rowcount
    
    # Delete area materials
    del_area = await session.execute(
        delete(ProjectAreaMaterial).where(ProjectAreaMaterial.project_id == project_id)
    )
    deleted_counts["area_materials"] = del_area.rowcount
    
    # Delete supply tracking
    del_supply = await session.execute(
        delete(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    deleted_counts["supply_tracking"] = del_supply.rowcount
    
    await session.commit()
    
    total_deleted = sum(deleted_counts.values())
    
    return {
        "message": f"تم حذف كميات المشروع بنجاح ({total_deleted} سجل)",
        "details": deleted_counts
    }


@router.post("/projects/{project_id}/enable")
async def enable_project_for_buildings(
    project_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Enable project for buildings system"""
    result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Set is_building_project to True
    project.is_building_project = True
    await session.commit()
    
    return {"message": "تم تفعيل المشروع في نظام الكميات بنجاح"}


@router.get("/my-supply-tracking")
async def get_my_supply_tracking(
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    Get supply tracking for projects assigned to the current user (supervisor/engineer)
    تتبع التوريد للمشاريع المرتبطة بالمستخدم الحالي فقط
    Returns empty list if no projects assigned
    """
    from database.models import SupplyTracking
    
    user_id = str(current_user.id)
    user_role = current_user.role
    
    # Build query based on user role - ONLY assigned projects
    if user_role == "supervisor":
        # Get projects where user is supervisor
        projects_query = select(Project).where(
            Project.supervisor_id == user_id,
            Project.status == "active"
        )
    elif user_role == "engineer":
        # Get projects where user is engineer
        projects_query = select(Project).where(
            Project.engineer_id == user_id,
            Project.status == "active"
        )
    else:
        # For other roles (procurement, admin), show all active projects
        projects_query = select(Project).where(Project.status == "active")
    
    projects_result = await session.execute(projects_query.order_by(Project.created_at.desc()))
    projects = projects_result.scalars().all()
    
    # Get supply tracking for each project
    result = []
    for project in projects:
        supply_result = await session.execute(
            select(SupplyTracking).where(SupplyTracking.project_id == str(project.id))
        )
        supply_items = supply_result.scalars().all()
        
        # Calculate totals
        total_required = sum(item.required_quantity for item in supply_items)
        total_received = sum(item.received_quantity or 0 for item in supply_items)
        
        project_data = {
            "project_id": str(project.id),
            "project_name": project.name,
            "project_code": project.code,
            "items_count": len(supply_items),
            "total_required": total_required,
            "total_received": total_received,
            "completion_percentage": round((total_received / total_required * 100) if total_required > 0 else 0, 1),
            "supply_items": [
                {
                    "id": str(item.id),
                    "item_name": item.item_name,
                    "item_code": item.item_code,
                    "unit": item.unit,
                    "required_quantity": item.required_quantity,
                    "received_quantity": item.received_quantity or 0,
                    "remaining_quantity": item.required_quantity - (item.received_quantity or 0),
                    "completion_percentage": round(((item.received_quantity or 0) / item.required_quantity * 100) if item.required_quantity > 0 else 0, 1)
                }
                for item in supply_items
            ]
        }
        result.append(project_data)
    
    return result


@router.get("/dashboard")
async def get_buildings_dashboard(
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get buildings system dashboard"""
    # Get projects count
    projects_result = await session.execute(select(func.count(Project.id)))
    total_projects = projects_result.scalar_one()
    
    # Get active projects
    active_result = await session.execute(
        select(func.count(Project.id)).where(Project.status == "active")
    )
    active_projects = active_result.scalar_one()
    
    return {
        "total_projects": total_projects,
        "active_projects": active_projects,
        "completed_projects": total_projects - active_projects,
        "templates_count": 0,  # To be implemented
        "total_floors": 0,  # To be implemented
        "recent_activity": []
    }


@router.get("/reports/supply-details/{project_id}")
async def get_supply_details_report(
    project_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get detailed supply report for a project"""
    from database.models import SupplyTracking
    
    # Get project
    result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Get supply tracking items for this project
    supply_result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    supply_items = supply_result.scalars().all()
    
    # Categorize items
    completed_items = []
    in_progress_items = []
    not_started_items = []
    
    total_required = 0
    total_received = 0
    total_required_value = 0
    total_received_value = 0
    
    for item in supply_items:
        required_qty = item.required_quantity or 0
        received_qty = item.received_quantity or 0
        unit_price = item.unit_price if hasattr(item, 'unit_price') else 0
        remaining = required_qty - received_qty
        completion = round((received_qty / required_qty * 100), 1) if required_qty > 0 else 0
        
        total_required += required_qty
        total_received += received_qty
        total_required_value += required_qty * unit_price
        total_received_value += received_qty * unit_price
        
        item_data = {
            "id": str(item.id),
            "item_code": getattr(item, 'item_code', None),
            "item_name": item.item_name,
            "unit": item.unit,
            "required_quantity": required_qty,
            "received_quantity": received_qty,
            "remaining_quantity": remaining,
            "completion_percentage": completion,
            "remaining_value": remaining * unit_price
        }
        
        if completion >= 100:
            completed_items.append(item_data)
        elif received_qty > 0:
            in_progress_items.append(item_data)
        else:
            not_started_items.append(item_data)
    
    total_items = len(supply_items)
    overall_completion = round((total_received / total_required * 100), 1) if total_required > 0 else 0
    
    # Return report data structure matching frontend expectations
    return {
        "project_id": project_id,
        "project_name": project.name,
        "summary": {
            "total_items": total_items,
            "completed_count": len(completed_items),
            "in_progress_count": len(in_progress_items),
            "not_started_count": len(not_started_items),
            "overall_completion": overall_completion,
            "total_required": total_required,
            "total_received": total_received,
            "total_remaining": total_required - total_received,
            "total_required_value": total_required_value,
            "total_received_value": total_received_value
        },
        "completed_items": completed_items,
        "in_progress_items": in_progress_items,
        "not_started_items": not_started_items,
        "materials": [],
        "delivery_timeline": [],
        "suppliers": []
    }


@router.get("/reports/summary")
async def get_buildings_reports_summary(
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get buildings reports summary"""
    # Get projects
    result = await session.execute(select(Project))
    projects = result.scalars().all()
    
    return {
        "total_projects": len(projects),
        "projects_by_status": {
            "active": len([p for p in projects if p.status == "active"]),
            "completed": len([p for p in projects if p.status == "completed"]),
            "on_hold": len([p for p in projects if p.status == "on_hold"])
        },
        "summary": {
            "total_area": 0,
            "total_materials": 0,
            "total_cost": 0
        }
    }


@router.get("/permissions/my")
async def get_my_permissions(
    current_user = Depends(get_current_user)
):
    """Get current user permissions for buildings system"""
    # Basic role-based permissions
    permissions = {
        "can_view": True,
        "can_edit": current_user.role in ["quantity_engineer", "system_admin", "procurement_manager"],
        "can_delete": current_user.role in ["system_admin", "procurement_manager"],
        "can_create_template": current_user.role in ["quantity_engineer", "system_admin"],
        "can_manage_floors": current_user.role in ["quantity_engineer", "system_admin"],
        "can_export": True,
        "assigned_projects": []  # To be implemented - project-specific permissions
    }
    
    return permissions


@router.get("/users/available")
async def get_available_users(
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get all users available for permission assignment"""
    from database.models import User
    
    result = await session.execute(
        select(User).where(User.is_active == True)
    )
    users = result.scalars().all()
    
    return [
        {
            "id": str(u.id),
            "name": u.name,
            "email": u.email,
            "role": u.role
        }
        for u in users
    ]


@router.get("/permissions")
async def get_all_permissions(
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get all building permissions (admin only)"""
    from database.models import BuildingsPermission, User
    
    # Check if table exists
    try:
        result = await session.execute(
            select(BuildingsPermission)
        )
        permissions = result.scalars().all()
        
        return [
            {
                "id": str(p.id),
                "user_id": p.user_id,
                "user_name": p.user_name,
                "project_id": p.project_id,
                "project_name": getattr(p, 'project_name', 'جميع المشاريع'),
                "can_view": p.can_view,
                "can_edit": p.can_edit,
                "can_delete": p.can_delete,
                "can_export": p.can_export
            }
            for p in permissions
        ]
    except Exception as e:
        print(f"Error fetching permissions: {e}")
        return []


@router.post("/permissions", status_code=status.HTTP_201_CREATED)
async def grant_permission(
    data: dict,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Grant building permission to user"""
    from database.models import BuildingsPermission, User, Project
    import uuid as uuid_lib
    
    user_id = data.get("user_id")
    
    # Get user info
    user_result = await session.execute(
        select(User).where(User.id == user_id)
    )
    user = user_result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # Get project info if specified
    project_id = data.get("project_id") or None
    project_name = "جميع المشاريع"
    if project_id:
        proj_result = await session.execute(
            select(Project).where(Project.id == project_id)
        )
        project = proj_result.scalar_one_or_none()
        if project:
            project_name = project.name
    
    # Create permission
    permission = BuildingsPermission(
        id=str(uuid_lib.uuid4()),
        user_id=str(user.id),
        user_name=user.name,
        project_id=project_id,
        project_name=project_name,
        can_view=data.get("can_view", True),
        can_edit=data.get("can_edit", False),
        can_delete=data.get("can_delete", False),
        can_export=data.get("can_export", True),
        granted_by=str(current_user.id),
        granted_by_name=current_user.name
    )
    
    session.add(permission)
    await session.commit()
    
    return {"message": "تم إعطاء الصلاحية بنجاح", "id": permission.id}


@router.delete("/permissions/{permission_id}")
async def revoke_permission(
    permission_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Revoke building permission"""
    from database.models import BuildingsPermission
    
    result = await session.execute(
        select(BuildingsPermission).where(BuildingsPermission.id == permission_id)
    )
    permission = result.scalar_one_or_none()
    
    if not permission:
        raise HTTPException(status_code=404, detail="الصلاحية غير موجودة")
    
    await session.delete(permission)
    await session.commit()
    
    return {"message": "تم إلغاء الصلاحية"}


# ==================== TEMPLATES ====================

@router.get("/projects/{project_id}/templates")
async def get_templates(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Get all templates for a project with materials
    Uses: BuildingsService -> BuildingsRepository
    """
    return await buildings_service.get_templates_by_project(project_id)


@router.post("/projects/{project_id}/templates", status_code=status.HTTP_201_CREATED)
async def create_template(
    project_id: str,
    data: TemplateCreate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    Create unit template
    Uses: BuildingsService -> BuildingsRepository
    """
    # Get project name
    result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    template = await buildings_service.create_template(
        project_id=project_id,
        project_name=project.name,
        code=data.code,
        name=data.name,
        area=data.area,
        rooms_count=data.rooms_count,
        bathrooms_count=data.bathrooms_count,
        count=data.count,
        description=data.description,
        created_by=str(current_user.id),
        created_by_name=current_user.name
    )
    return {
        "id": template.id,
        "code": template.code,
        "name": template.name,
        "message": "تم إنشاء النموذج بنجاح"
    }


@router.put("/templates/{template_id}")
async def update_template(
    template_id: str,
    data: TemplateUpdate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Update unit template
    Uses: BuildingsService -> BuildingsRepository
    """
    update_data = data.model_dump(exclude_unset=True)
    template = await buildings_service.update_template(template_id, update_data)
    
    if not template:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="النموذج غير موجود"
        )
    return {"message": "تم تحديث النموذج بنجاح"}


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Delete unit template
    Uses: BuildingsService -> BuildingsRepository
    """
    success = await buildings_service.delete_template(template_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="النموذج غير موجود"
        )
    return {"message": "تم حذف النموذج بنجاح"}


@router.post("/templates/{template_id}/materials", status_code=status.HTTP_201_CREATED)
async def add_template_material(
    template_id: str,
    data: TemplateMaterialCreate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Add material to template
    Uses: BuildingsService -> BuildingsRepository
    """
    material = await buildings_service.add_template_material(
        template_id=template_id,
        catalog_item_id=data.catalog_item_id,
        item_code=data.item_code,
        item_name=data.item_name,
        unit=data.unit,
        quantity_per_unit=data.quantity_per_unit,
        unit_price=data.unit_price
    )
    return {"id": material.id, "message": "تم إضافة المادة بنجاح"}


@router.delete("/templates/{template_id}/materials/{material_id}")
async def delete_template_material(
    template_id: str,
    material_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Delete material from template
    Uses: BuildingsService -> BuildingsRepository
    """
    success = await buildings_service.delete_template_material(material_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="المادة غير موجودة"
        )
    return {"message": "تم حذف المادة بنجاح"}


# ==================== FLOORS ====================

@router.get("/projects/{project_id}/floors")
async def get_floors(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Get all floors for a project
    Uses: BuildingsService -> BuildingsRepository
    """
    floors = await buildings_service.get_floors_by_project(project_id)
    return [floor_to_response(f) for f in floors]


@router.post("/projects/{project_id}/floors", status_code=status.HTTP_201_CREATED)
async def create_floor(
    project_id: str,
    data: FloorCreate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Create project floor
    Uses: BuildingsService -> BuildingsRepository
    """
    floor = await buildings_service.create_floor(
        project_id=project_id,
        floor_number=data.floor_number,
        floor_name=data.floor_name,
        area=data.area,
        steel_factor=data.steel_factor,
        created_by=str(current_user.id),
        created_by_name=current_user.name
    )
    return floor_to_response(floor)


@router.put("/projects/{project_id}/floors/{floor_id}")
async def update_floor(
    project_id: str,
    floor_id: str,
    data: FloorUpdate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Update project floor
    Uses: BuildingsService -> BuildingsRepository
    """
    update_data = data.model_dump(exclude_unset=True)
    floor = await buildings_service.update_floor(floor_id, update_data)
    
    if not floor:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الدور غير موجود"
        )
    return floor_to_response(floor)


@router.delete("/projects/{project_id}/floors/{floor_id}")
async def delete_floor(
    project_id: str,
    floor_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Delete project floor
    Uses: BuildingsService -> BuildingsRepository
    """
    success = await buildings_service.delete_floor(floor_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="الدور غير موجود"
        )
    return {"message": "تم حذف الدور بنجاح"}


# ==================== AREA MATERIALS ====================

@router.get("/projects/{project_id}/area-materials")
async def get_area_materials(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Get area materials for a project
    Uses: BuildingsService -> BuildingsRepository
    """
    materials = await buildings_service.get_area_materials_by_project(project_id)
    return [area_material_to_response(m) for m in materials]


@router.post("/projects/{project_id}/area-materials", status_code=status.HTTP_201_CREATED)
async def create_area_material(
    project_id: str,
    data: AreaMaterialCreate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Create area material with advanced options
    Uses: BuildingsService -> BuildingsRepository
    """
    material = await buildings_service.create_area_material(
        project_id=project_id,
        catalog_item_id=data.catalog_item_id,
        item_name=data.item_name,
        unit=data.unit,
        calculation_method=data.calculation_method,
        factor=data.factor,
        direct_quantity=data.direct_quantity,
        unit_price=data.unit_price,
        calculation_type=data.calculation_type,
        selected_floor_id=data.selected_floor_id,
        tile_width=data.tile_width,
        tile_height=data.tile_height,
        waste_percentage=data.waste_percentage,
        notes=data.notes,
        created_by=str(current_user.id),
        created_by_name=current_user.name
    )
    return area_material_to_response(material)


@router.put("/projects/{project_id}/area-materials/{material_id}")
async def update_area_material(
    project_id: str,
    material_id: str,
    data: AreaMaterialUpdate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Update area material
    Uses: BuildingsService -> BuildingsRepository
    """
    update_data = data.model_dump(exclude_unset=True)
    material = await buildings_service.update_area_material(material_id, update_data)
    
    if not material:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="المادة غير موجودة"
        )
    return area_material_to_response(material)


@router.delete("/projects/{project_id}/area-materials/{material_id}")
async def delete_area_material(
    project_id: str,
    material_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Delete area material
    Uses: BuildingsService -> BuildingsRepository
    """
    success = await buildings_service.delete_area_material(material_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="المادة غير موجودة"
        )
    return {"message": "تم حذف المادة بنجاح"}


# ==================== SUPPLY TRACKING ====================

@router.get("/projects/{project_id}/supply")
async def get_supply(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Get supply tracking for a project
    Uses: BuildingsService -> BuildingsRepository
    """
    items = await buildings_service.get_supply_by_project(project_id)
    return [supply_to_response(item) for item in items]


@router.post("/projects/{project_id}/sync-supply")
async def sync_supply_tracking(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    Sync supply tracking with calculated quantities
    Creates supply items for all calculated materials
    """
    from database.models import SupplyTracking, UnitTemplate, UnitTemplateMaterial, ProjectFloor, ProjectAreaMaterial
    
    # Get project
    result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Delete existing supply items for this project
    await session.execute(
        delete(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    
    # Get templates with materials
    templates_result = await session.execute(
        select(UnitTemplate).where(UnitTemplate.project_id == project_id)
    )
    templates = templates_result.scalars().all()
    
    # Get floors for total area
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id)
    )
    floors = floors_result.scalars().all()
    total_area = sum(f.area for f in floors)
    
    # Get area materials
    area_materials_result = await session.execute(
        select(ProjectAreaMaterial).where(ProjectAreaMaterial.project_id == project_id)
    )
    area_materials = area_materials_result.scalars().all()
    
    created_items = []
    
    # Create supply items from template materials
    for template in templates:
        materials_result = await session.execute(
            select(UnitTemplateMaterial).where(UnitTemplateMaterial.template_id == template.id)
        )
        materials = materials_result.scalars().all()
        
        for m in materials:
            quantity = m.quantity_per_unit * template.count
            supply_item = SupplyTracking(
                id=str(uuid4()),
                project_id=project_id,
                catalog_item_id=m.catalog_item_id,
                item_code=m.item_code,
                item_name=f"{m.item_name} ({template.name})",
                unit=m.unit,
                required_quantity=quantity,
                received_quantity=0,
                unit_price=m.unit_price,
                source="quantity",
                notes=f"من نموذج: {template.name}"
            )
            session.add(supply_item)
            created_items.append(supply_item)
    
    # Create supply items from area materials
    for m in area_materials:
        calc_method = getattr(m, 'calculation_method', 'factor') or 'factor'
        
        if calc_method == 'direct':
            base_quantity = getattr(m, 'direct_quantity', 0) or 0
            floor_area = 0
        else:
            # Get floor area
            if getattr(m, 'calculation_type', 'all_floors') == 'selected_floor' and getattr(m, 'selected_floor_id', None):
                floor_area = next((f.area for f in floors if str(f.id) == m.selected_floor_id), 0)
            else:
                floor_area = total_area
            base_quantity = floor_area * m.factor
        
        # Handle tile calculation
        tile_width = getattr(m, 'tile_width', 0) or 0
        tile_height = getattr(m, 'tile_height', 0) or 0
        if tile_width > 0 and tile_height > 0 and floor_area > 0:
            tile_area_m2 = (tile_width / 100) * (tile_height / 100)
            if tile_area_m2 > 0:
                base_quantity = floor_area / tile_area_m2
        
        # Apply waste percentage
        waste_pct = getattr(m, 'waste_percentage', 0) or 0
        quantity = base_quantity * (1 + waste_pct / 100)
        
        supply_item = SupplyTracking(
            id=str(uuid4()),
            project_id=project_id,
            catalog_item_id=m.catalog_item_id,
            item_code=getattr(m, 'item_code', None),
            item_name=m.item_name,
            unit=m.unit,
            required_quantity=round(quantity, 2),
            received_quantity=0,
            unit_price=m.unit_price,
            source="area",
            notes="مواد المساحة"
        )
        session.add(supply_item)
        created_items.append(supply_item)
    
    await session.commit()
    
    return {
        "message": f"تم مزامنة التوريد بنجاح ({len(created_items)} عنصر)",
        "items_count": len(created_items)
    }


@router.put("/projects/{project_id}/supply/{item_id}")
async def update_supply_item(
    project_id: str,
    item_id: str,
    data: SupplyUpdate,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Update supply item
    Uses: BuildingsService -> BuildingsRepository
    """
    update_data = data.model_dump(exclude_unset=True)
    item = await buildings_service.update_supply_item(item_id, update_data)
    
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="العنصر غير موجود"
        )
    return supply_to_response(item)


# ==================== CALCULATIONS ====================

@router.get("/projects/{project_id}/calculate")
async def calculate_quantities(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Calculate all quantities for a project (BOQ)
    Uses: BuildingsService -> BuildingsRepository
    """
    return await buildings_service.calculate_project_quantities(project_id)


@router.get("/projects/{project_id}/stats")
async def get_project_stats(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service)
):
    """
    Get project statistics
    Uses: BuildingsService -> BuildingsRepository
    """
    return await buildings_service.get_project_stats(project_id)


# ==================== EXPORT / IMPORT ====================

@router.get("/projects/{project_id}/export/boq-excel")
async def export_boq_excel(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Export BOQ to Excel file"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from database.models import UnitTemplate, UnitTemplateMaterial, ProjectFloor, ProjectAreaMaterial
    
    # Get project
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Calculate quantities
    calc_data = await buildings_service.calculate_project_quantities(project_id)
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # === Sheet 1: Summary ===
    ws = wb.active
    ws.title = "ملخص"
    ws.sheet_view.rightToLeft = True
    
    ws['A1'] = f"جدول الكميات - {project.name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    
    ws['A3'] = "المشروع:"
    ws['B3'] = project.name
    ws['A4'] = "المالك:"
    ws['B4'] = project.owner_name or ""
    ws['A5'] = "إجمالي المساحة:"
    ws['B5'] = f"{calc_data['total_area']} م²"
    ws['A6'] = "إجمالي الوحدات:"
    ws['B6'] = calc_data['total_units']
    ws['A7'] = "إجمالي الحديد:"
    ws['B7'] = f"{calc_data['steel_calculation']['total_steel_tons']} طن"
    ws['A8'] = "إجمالي التكلفة:"
    ws['B8'] = f"{calc_data['total_materials_cost']:,.2f} ر.س"
    
    # === Sheet 2: Steel by Floor ===
    ws2 = wb.create_sheet("الحديد حسب الدور")
    ws2.sheet_view.rightToLeft = True
    
    headers = ["الدور", "المساحة (م²)", "المعامل (كجم/م²)", "الحديد (طن)"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    for row, floor in enumerate(calc_data['steel_calculation']['floors'], 2):
        ws2.cell(row=row, column=1, value=floor['floor_name']).border = border
        ws2.cell(row=row, column=2, value=floor['area']).border = border
        ws2.cell(row=row, column=3, value=floor['steel_factor']).border = border
        ws2.cell(row=row, column=4, value=floor['steel_tons']).border = border
    
    # === Sheet 3: Unit Materials ===
    ws3 = wb.create_sheet("مواد الوحدات")
    ws3.sheet_view.rightToLeft = True
    
    headers = ["الكود", "المادة", "الوحدة", "الكمية", "السعر", "الإجمالي"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    for row, mat in enumerate(calc_data['materials'], 2):
        ws3.cell(row=row, column=1, value=mat.get('item_code', '')).border = border
        ws3.cell(row=row, column=2, value=mat['item_name']).border = border
        ws3.cell(row=row, column=3, value=mat['unit']).border = border
        ws3.cell(row=row, column=4, value=mat['quantity']).border = border
        ws3.cell(row=row, column=5, value=mat['unit_price']).border = border
        ws3.cell(row=row, column=6, value=mat['total_price']).border = border
    
    # Total row
    total_row = len(calc_data['materials']) + 2
    ws3.cell(row=total_row, column=5, value="الإجمالي:").font = Font(bold=True)
    ws3.cell(row=total_row, column=6, value=calc_data['total_unit_materials_cost']).font = Font(bold=True)
    
    # === Sheet 4: Area Materials ===
    ws4 = wb.create_sheet("مواد المساحة")
    ws4.sheet_view.rightToLeft = True
    
    headers = ["المادة", "الوحدة", "المعامل", "الدور", "الهالك%", "الكمية", "السعر", "الإجمالي"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    for row, mat in enumerate(calc_data['area_materials'], 2):
        ws4.cell(row=row, column=1, value=mat['item_name']).border = border
        ws4.cell(row=row, column=2, value=mat['unit']).border = border
        ws4.cell(row=row, column=3, value=mat.get('factor', 0)).border = border
        ws4.cell(row=row, column=4, value=mat.get('floor_name', 'جميع الأدوار')).border = border
        ws4.cell(row=row, column=5, value=mat.get('waste_percentage', 0)).border = border
        ws4.cell(row=row, column=6, value=mat['quantity']).border = border
        ws4.cell(row=row, column=7, value=mat['unit_price']).border = border
        ws4.cell(row=row, column=8, value=mat['total_price']).border = border
    
    # Total row
    total_row = len(calc_data['area_materials']) + 2
    ws4.cell(row=total_row, column=7, value="الإجمالي:").font = Font(bold=True)
    ws4.cell(row=total_row, column=8, value=calc_data['total_area_materials_cost']).font = Font(bold=True)
    
    # Adjust column widths - use column index to avoid MergedCell issues
    from openpyxl.utils import get_column_letter
    
    for ws in [wb.active, ws2, ws3, ws4]:
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Skip merged cells
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            ws.column_dimensions[column_letter].width = max(max_length + 5, 12)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from urllib.parse import quote
    
    # Use URL encoding for Arabic filename
    safe_filename = f"BOQ_{project_id[:8]}.xlsx"
    encoded_filename = quote(f"BOQ_{project.name}.xlsx")
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.get("/projects/{project_id}/export/boq-pdf")
async def export_boq_pdf(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Export BOQ to PDF file with Arabic support"""
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from io import BytesIO
    import os
    import arabic_reshaper
    from bidi.algorithm import get_display
    
    # Helper function for Arabic text
    def arabic(text):
        if not text:
            return ""
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)
        except:
            return str(text)
    
    # Register Arabic font
    font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('Arabic', font_path))
        except:
            pass
    
    # Get project
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Calculate quantities
    calc_data = await buildings_service.calculate_project_quantities(project_id)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    
    # Arabic styles
    title_style = ParagraphStyle('ArabicTitle', fontName='Arabic', fontSize=18, alignment=TA_CENTER, spaceAfter=20)
    heading_style = ParagraphStyle('ArabicHeading', fontName='Arabic', fontSize=14, alignment=TA_RIGHT, spaceAfter=10, spaceBefore=20)
    
    # Title
    elements.append(Paragraph(arabic(f"جدول الكميات - {project.name}"), title_style))
    elements.append(Spacer(1, 20))
    
    # Summary table
    summary_data = [
        [arabic("القيمة"), arabic("البيان")],
        [project.name, arabic("المشروع")],
        [project.owner_name or "-", arabic("المالك")],
        [f"{calc_data['total_area']} م²", arabic("إجمالي المساحة")],
        [str(calc_data['total_units']), arabic("عدد الوحدات")],
        [f"{calc_data['steel_calculation']['total_steel_tons']} طن", arabic("إجمالي الحديد")],
        [f"{calc_data['total_materials_cost']:,.2f} ريال", arabic("إجمالي التكلفة")],
    ]
    
    summary_table = Table(summary_data, colWidths=[250, 200])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('BACKGROUND', (1, 0), (1, 0), colors.Color(0.18, 0.49, 0.2)),
        ('BACKGROUND', (0, 0), (0, 0), colors.Color(0.18, 0.49, 0.2)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Steel table
    elements.append(Paragraph(arabic("الحديد حسب الدور"), heading_style))
    steel_data = [[arabic("الحديد (طن)"), arabic("المعامل (كجم/م²)"), arabic("المساحة (م²)"), arabic("الدور")]]
    for floor in calc_data['steel_calculation']['floors']:
        steel_data.append([
            str(floor['steel_tons']),
            str(floor['steel_factor']),
            str(floor['area']),
            arabic(floor['floor_name'])
        ])
    steel_data.append([
        str(calc_data['steel_calculation']['total_steel_tons']),
        "",
        str(calc_data['total_area']),
        arabic("الإجمالي")
    ])
    
    steel_table = Table(steel_data, colWidths=[100, 100, 100, 150])
    steel_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.18, 0.49, 0.2)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(steel_table)
    elements.append(Spacer(1, 30))
    
    # Unit Materials table
    if calc_data['materials']:
        elements.append(Paragraph(arabic("مواد الوحدات"), heading_style))
        mat_data = [[arabic("الإجمالي"), arabic("السعر"), arabic("الكمية"), arabic("الوحدة"), arabic("المادة"), arabic("الكود")]]
        for mat in calc_data['materials']:
            mat_data.append([
                f"{mat['total_price']:,.2f}",
                str(mat['unit_price']),
                str(mat['quantity']),
                arabic(mat['unit']),
                arabic(mat['item_name'][:30]),
                mat.get('item_code', '')
            ])
        mat_data.append([f"{calc_data['total_unit_materials_cost']:,.2f}", arabic("الإجمالي:"), "", "", "", ""])
        
        mat_table = Table(mat_data, colWidths=[80, 60, 60, 50, 150, 60])
        mat_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.18, 0.49, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(mat_table)
        elements.append(Spacer(1, 30))
    
    # Area Materials table
    if calc_data['area_materials']:
        elements.append(Paragraph(arabic("مواد المساحة"), heading_style))
        area_data = [[arabic("الإجمالي"), arabic("السعر"), arabic("الكمية"), arabic("الهالك%"), arabic("المعامل"), arabic("الوحدة"), arabic("المادة")]]
        for mat in calc_data['area_materials']:
            area_data.append([
                f"{mat['total_price']:,.2f}",
                str(mat['unit_price']),
                str(mat['quantity']),
                str(mat.get('waste_percentage', 0)),
                str(mat.get('factor', 0)),
                arabic(mat['unit']),
                arabic(mat['item_name'][:25])
            ])
        area_data.append([f"{calc_data['total_area_materials_cost']:,.2f}", arabic("الإجمالي:"), "", "", "", "", ""])
        
        area_table = Table(area_data, colWidths=[80, 60, 60, 50, 50, 50, 110])
        area_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.18, 0.49, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(area_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    from urllib.parse import quote
    safe_filename = f"BOQ_{project_id[:8]}.pdf"
    encoded_filename = quote(f"جدول_كميات_{project.name}.pdf")
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.get("/projects/{project_id}/export/floors-excel")
async def export_floors_excel(
    project_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Export floors to Excel file"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from database.models import ProjectFloor
    
    # Get project
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Get floors
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id).order_by(ProjectFloor.floor_number)
    )
    floors = floors_result.scalars().all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "الأدوار"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["رقم الدور", "اسم الدور", "المساحة (م²)", "معامل التسليح (كجم/م²)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Data
    for row, floor in enumerate(floors, 2):
        ws.cell(row=row, column=1, value=floor.floor_number).border = border
        ws.cell(row=row, column=2, value=floor.floor_name).border = border
        ws.cell(row=row, column=3, value=floor.area).border = border
        ws.cell(row=row, column=4, value=floor.steel_factor).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from urllib.parse import quote
    safe_filename = f"Floors_{project_id[:8]}.xlsx"
    encoded_filename = quote(f"Floors_{project.name}.xlsx")
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.post("/projects/{project_id}/import/floors")
async def import_floors_excel(
    project_id: str,
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Import floors from Excel file"""
    from openpyxl import load_workbook
    from io import BytesIO
    from database.models import ProjectFloor
    
    # Get project
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Read Excel file
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    
    imported_count = 0
    errors = []
    
    # Skip header row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] and not row[1]:
            continue
        
        try:
            floor_number = int(row[0]) if row[0] else 0
            floor_name = str(row[1]) if row[1] else f"الدور {floor_number}"
            area = float(row[2]) if row[2] else 0
            steel_factor = float(row[3]) if row[3] else 120
            
            floor = ProjectFloor(
                id=str(uuid4()),
                project_id=project_id,
                floor_number=floor_number,
                floor_name=floor_name,
                area=area,
                steel_factor=steel_factor
            )
            session.add(floor)
            imported_count += 1
        except Exception as e:
            errors.append(f"Row {row}: {str(e)}")
    
    await session.commit()
    
    return {
        "message": f"تم استيراد {imported_count} دور بنجاح",
        "imported_count": imported_count,
        "errors": errors if errors else None
    }


@router.get("/export/project-template")
async def download_project_template(current_user = Depends(get_current_user)):
    """Download project import template (Excel)"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from io import BytesIO
    
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    note_font = Font(italic=True, color="666666")
    
    # ==================== Sheet 1: الأدوار ====================
    ws1 = wb.active
    ws1.title = "الأدوار"
    ws1.sheet_view.rightToLeft = True
    
    headers = ["رقم الدور", "اسم الدور", "المساحة (م²)", "معامل التسليح (كجم/م²)"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Sample data
    sample_floors = [
        [0, "اللبشة", 430, 120],
        [1, "دور المواقف", 800, 120],
        [2, "الدور الأول", 300, 120],
        [3, "الدور الثاني", 300, 120],
    ]
    for row, data in enumerate(sample_floors, 2):
        for col, value in enumerate(data, 1):
            ws1.cell(row=row, column=col, value=value)
    
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 25
    
    # ==================== Sheet 2: النماذج ====================
    ws2 = wb.create_sheet("النماذج")
    ws2.sheet_view.rightToLeft = True
    
    headers = ["الكود", "الاسم", "عدد الوحدات", "المساحة (م²)", "الغرف", "الحمامات"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    sample_templates = [
        ["A", "شقة أمامية", 4, 200, 4, 3],
        ["B", "شقة خلفية", 3, 150, 3, 2],
    ]
    for row, data in enumerate(sample_templates, 2):
        for col, value in enumerate(data, 1):
            ws2.cell(row=row, column=col, value=value)
    
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    
    # ==================== Sheet 3: مواد المساحة ====================
    ws3 = wb.create_sheet("مواد المساحة")
    ws3.sheet_view.rightToLeft = True
    
    # نفس الأعمدة المستخدمة في التصدير
    headers = ["اسم المادة", "الوحدة", "المعامل", "الدور", "نسبة الهالك %", "الكمية", "السعر", "الإجمالي"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # ملاحظات
    ws3.cell(row=2, column=1, value="# ملاحظة: اترك الدور فارغاً لتطبيق المادة على جميع الأدوار. الكمية المدخلة هي الكمية المباشرة.")
    ws3.cell(row=2, column=1).font = note_font
    ws3.merge_cells('A2:H2')
    
    sample_materials = [
        ["حديد 8 ملم", "طن", 0, "صبة الارضيه", 0, 2, 0, 0],
        ["حديد 10 ملم", "طن", 0, "اعمدة الارضي", 0, 2.7, 0, 0],
        ["خرسانه c15", "م³", 0, "صبة الارضيه", 0, 49, 0, 0],
        ["خرسانه c35", "م³", 0, "اعمدة الارضي", 0, 33, 0, 0],
        ["حديد 12 ملم", "طن", 0, "سقف الارضي", 0, 15.9, 0, 0],
    ]
    for row, data in enumerate(sample_materials, 3):
        for col, value in enumerate(data, 1):
            ws3.cell(row=row, column=col, value=value)
    
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 15
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 12
    ws3.column_dimensions['H'].width = 12
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=project_template.xlsx"}
    )


@router.post("/projects/{project_id}/import/area-materials")
async def import_area_materials_excel(
    project_id: str,
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Import area materials from Excel file - استيراد مواد المساحة من Excel"""
    from openpyxl import load_workbook
    from io import BytesIO
    from database.models import ProjectFloor, ProjectAreaMaterial
    
    # Get project
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Get floors for mapping (by number and by name)
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id)
    )
    floors = floors_result.scalars().all()
    floors_by_number = {f.floor_number: f for f in floors}
    floors_by_name = {f.floor_name.strip(): f for f in floors}
    
    # Read Excel file
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    
    imported_count = 0
    errors = []
    
    # Detect format from header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    # Check if column 6 is "الكمية" (quantity) - means direct quantity format
    is_quantity_format = header_row and len(header_row) > 5 and header_row[5] and "كمية" in str(header_row[5])
    
    # Skip header row (and note row if exists)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows or note rows
        if not row[0] or str(row[0]).startswith('#'):
            continue
        
        try:
            item_name = str(row[0]).strip()
            unit = str(row[1]).strip() if row[1] else "طن"
            factor = float(row[2]) if row[2] else 0
            floor_value = row[3] if len(row) > 3 else None
            waste_percentage = float(row[4]) if len(row) > 4 and row[4] else 0
            
            # Handle different formats
            if is_quantity_format:
                # Format: اسم المادة, الوحدة, المعامل, الدور (اسم), نسبة الهالك, الكمية, السعر, الإجمالي
                direct_quantity = float(row[5]) if len(row) > 5 and row[5] else 0
                unit_price = float(row[6]) if len(row) > 6 and row[6] else 0
                calculation_method = "direct" if direct_quantity > 0 else "factor"
            else:
                # Format: اسم المادة, الوحدة, المعامل, رقم الدور, نسبة الهالك, السعر
                direct_quantity = 0
                unit_price = float(row[5]) if len(row) > 5 and row[5] else 0
                calculation_method = "factor"
            
            # Determine floor
            calculation_type = "all_floors"
            selected_floor_id = None
            
            if floor_value not in [None, "", " "]:
                floor_str = str(floor_value).strip()
                # Try by name first
                if floor_str in floors_by_name:
                    calculation_type = "selected_floor"
                    selected_floor_id = str(floors_by_name[floor_str].id)
                else:
                    # Try by number
                    try:
                        floor_num = int(floor_value)
                        if floor_num in floors_by_number:
                            calculation_type = "selected_floor"
                            selected_floor_id = str(floors_by_number[floor_num].id)
                    except (ValueError, TypeError):
                        # Floor name not found, keep as all_floors
                        pass
            
            material = ProjectAreaMaterial(
                id=str(uuid4()),
                project_id=project_id,
                catalog_item_id=None,
                item_name=item_name,
                unit=unit,
                calculation_method=calculation_method,
                factor=factor,
                direct_quantity=direct_quantity,
                unit_price=unit_price,
                calculation_type=calculation_type,
                selected_floor_id=selected_floor_id,
                tile_width=0,
                tile_height=0,
                waste_percentage=waste_percentage,
                notes=None
            )
            session.add(material)
            imported_count += 1
        except Exception as e:
            errors.append(f"صف {row_idx}: {str(e)}")
    
    await session.commit()
    
    return {
        "message": f"تم استيراد {imported_count} مادة بنجاح",
        "imported_count": imported_count,
        "errors": errors if errors else None
    }


@router.get("/projects/{project_id}/export/area-materials")
async def export_area_materials_excel(
    project_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Export area materials to Excel file - تصدير مواد المساحة إلى Excel (متوافق مع الاستيراد)"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from io import BytesIO
    from database.models import ProjectFloor, ProjectAreaMaterial
    
    # Get project
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Get floors
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id).order_by(ProjectFloor.floor_number)
    )
    floors = floors_result.scalars().all()
    floors_dict = {str(f.id): f for f in floors}
    total_area = sum(f.area for f in floors)
    
    # Get area materials
    materials_result = await session.execute(
        select(ProjectAreaMaterial).where(ProjectAreaMaterial.project_id == project_id)
    )
    area_materials = materials_result.scalars().all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "مواد المساحة"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers - متطابقة مع تنسيق الاستيراد
    headers = ["اسم المادة", "الوحدة", "المعامل", "الدور", "نسبة الهالك %", "الكمية", "السعر", "الإجمالي"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Data
    row = 2
    total_cost = 0
    for mat in area_materials:
        calc_type = getattr(mat, 'calculation_type', 'all_floors') or 'all_floors'
        calc_method = getattr(mat, 'calculation_method', 'factor') or 'factor'
        
        # اسم الدور
        floor_name = ""  # فارغ = جميع الأدوار
        floor_area = total_area
        
        if calc_type == 'selected_floor' and getattr(mat, 'selected_floor_id', None):
            floor = floors_dict.get(mat.selected_floor_id)
            if floor:
                floor_name = floor.floor_name
                floor_area = floor.area
        
        # حساب الكمية
        if calc_method == 'direct':
            quantity = getattr(mat, 'direct_quantity', 0) or 0
            factor_value = 0  # المعامل = 0 للكمية المباشرة
        else:
            factor_value = mat.factor or 0
            quantity = floor_area * factor_value
        
        # حساب البلاط
        tile_width = getattr(mat, 'tile_width', 0) or 0
        tile_height = getattr(mat, 'tile_height', 0) or 0
        if tile_width > 0 and tile_height > 0 and floor_area > 0:
            tile_area_m2 = (tile_width / 100) * (tile_height / 100)
            if tile_area_m2 > 0:
                quantity = floor_area / tile_area_m2
        
        # تطبيق نسبة الهالك
        waste_pct = getattr(mat, 'waste_percentage', 0) or 0
        quantity_with_waste = quantity * (1 + waste_pct / 100)
        
        total_price = quantity_with_waste * (mat.unit_price or 0)
        total_cost += total_price
        
        ws.cell(row=row, column=1, value=mat.item_name).border = border
        ws.cell(row=row, column=2, value=mat.unit).border = border
        ws.cell(row=row, column=3, value=factor_value).border = border
        ws.cell(row=row, column=4, value=floor_name).border = border
        ws.cell(row=row, column=5, value=waste_pct).border = border
        ws.cell(row=row, column=6, value=round(quantity_with_waste, 2)).border = border
        ws.cell(row=row, column=7, value=mat.unit_price or 0).border = border
        ws.cell(row=row, column=8, value=round(total_price, 2)).border = border
        row += 1
    
    # Total row
    ws.cell(row=row, column=7, value="الإجمالي:").font = Font(bold=True)
    ws.cell(row=row, column=8, value=round(total_cost, 2)).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from urllib.parse import quote
    safe_filename = f"Area_Materials_{project_id[:8]}.xlsx"
    encoded_filename = quote(f"مواد_المساحة_{project.name}.xlsx")
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}"
        }
    )



@router.post("/supply/sync-from-delivery")
async def sync_supply_from_deliveries(
    project_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    مزامنة الكميات المستلمة من أوامر الشراء إلى نظام تتبع التوريد
    """
    from database.models import SupplyTracking, PurchaseOrder, PurchaseOrderItem
    
    # Get project
    result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Get all delivered order items for this project
    delivered_result = await session.execute(
        select(PurchaseOrderItem, PurchaseOrder)
        .join(PurchaseOrder, PurchaseOrderItem.order_id == PurchaseOrder.id)
        .where(
            PurchaseOrder.project_id == project_id,
            PurchaseOrder.status.in_(['delivered', 'partially_delivered']),
            PurchaseOrderItem.delivered_quantity > 0
        )
    )
    delivered_items = delivered_result.all()
    
    # Get all supply tracking items for this project
    supply_result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    supply_items = {item.item_name: item for item in supply_result.scalars().all()}
    
    # Reset all received quantities first
    for supply_item in supply_items.values():
        supply_item.received_quantity = 0
    
    # Update supply tracking based on delivered items
    updated_count = 0
    for poi, po in delivered_items:
        item_name = poi.name
        delivered_qty = poi.delivered_quantity or 0
        
        # Try exact match first
        if item_name in supply_items:
            supply_items[item_name].received_quantity += delivered_qty
            updated_count += 1
        else:
            # Try partial match
            for supply_name, supply_item in supply_items.items():
                if item_name in supply_name or supply_name in item_name:
                    supply_item.received_quantity += delivered_qty
                    updated_count += 1
                    break
    
    await session.commit()
    
    return {
        "message": "تمت المزامنة بنجاح",
        "project_id": project_id,
        "project_name": project.name,
        "items_synced": updated_count,
        "delivered_orders_count": len(delivered_items)
    }



@router.get("/reports/supply-export/{project_id}")
async def export_supply_report(
    project_id: str,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير تقرير التوريد إلى Excel"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    # Import models
    from database.models import SupplyTracking
    
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # جلب بيانات التوريد
    supply_result = await session.execute(
        select(SupplyTracking).where(SupplyTracking.project_id == project_id)
    )
    supply_items = supply_result.scalars().all()
    
    # إنشاء Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير التوريد"
    ws.sheet_view.rightToLeft = True
    
    # التنسيقات
    title_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    completed_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    progress_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    not_started_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    
    # العنوان
    ws.merge_cells('A1:I1')
    ws['A1'] = f'تقرير التوريد - {project.name}'
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # رؤوس الأعمدة
    headers = ['الكود', 'المادة', 'الوحدة', 'المطلوب', 'المستلم', 'المتبقي', 'الإنجاز %', 'السعر', 'القيمة المتبقية']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # البيانات
    row = 4
    for item in supply_items:
        required = item.required_quantity or 0
        received = item.received_quantity or 0
        remaining = required - received
        completion = (received / required * 100) if required > 0 else 0
        remaining_value = remaining * (item.unit_price or 0)
        
        # تحديد لون الصف
        if completion >= 100:
            row_fill = completed_fill
        elif completion > 0:
            row_fill = progress_fill
        else:
            row_fill = not_started_fill
        
        data = [
            item.item_code or "",
            item.item_name,
            item.unit,
            required,
            received,
            remaining,
            f"{completion:.1f}%",
            item.unit_price or 0,
            remaining_value
        ]
        
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
        
        row += 1
    
    # الإجماليات
    total_row = row
    ws.cell(row=total_row, column=2, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(s.required_quantity or 0 for s in supply_items)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(s.received_quantity or 0 for s in supply_items)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum((s.required_quantity or 0) - (s.received_quantity or 0) for s in supply_items)).font = Font(bold=True)
    
    total_required = sum(s.required_quantity or 0 for s in supply_items)
    total_received = sum(s.received_quantity or 0 for s in supply_items)
    overall_completion = (total_received / total_required * 100) if total_required > 0 else 0
    ws.cell(row=total_row, column=7, value=f"{overall_completion:.1f}%").font = Font(bold=True)
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    
    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from urllib.parse import quote
    filename = quote(f"Supply_Report_{project.name}.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/projects/{project_id}/export/materials-requests")
async def export_materials_requests(
    project_id: str,
    current_user = Depends(get_current_user),
    buildings_service: BuildingsService = Depends(get_buildings_service),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    تصدير طلبات المواد - فقط الأصناف التي فيها كميات مع توضيح الدور
    Export materials requests - only items with quantities, grouped by floor
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from database.models import ProjectFloor, ProjectAreaMaterial
    from datetime import datetime
    
    # Get project
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # Get floors
    floors_result = await session.execute(
        select(ProjectFloor).where(ProjectFloor.project_id == project_id).order_by(ProjectFloor.floor_number)
    )
    floors = floors_result.scalars().all()
    floors_dict = {str(f.id): f for f in floors}
    total_area = sum(f.area for f in floors)
    
    # Get area materials
    materials_result = await session.execute(
        select(ProjectAreaMaterial).where(ProjectAreaMaterial.project_id == project_id)
    )
    area_materials = materials_result.scalars().all()
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    subheader_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    data_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
    total_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    # ==================== Sheet 1: طلبات المواد حسب الدور ====================
    ws = wb.active
    ws.title = "طلبات المواد"
    ws.sheet_view.rightToLeft = True
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"طلبات المواد - {project.name}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    
    # Date
    ws.merge_cells('A2:H2')
    ws['A2'] = f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].alignment = center_align
    
    # Process materials grouped by floor
    row = 4
    
    # Group materials by calculation_type and floor
    materials_by_floor = {}  # {floor_id: [materials]}
    materials_all_floors = []  # Materials for all floors
    
    for m in area_materials:
        calc_type = getattr(m, 'calculation_type', 'all_floors') or 'all_floors'
        calc_method = getattr(m, 'calculation_method', 'factor') or 'factor'
        
        # Calculate quantity
        if calc_method == 'direct':
            base_qty = getattr(m, 'direct_quantity', 0) or 0
            floor_area = 0
        else:
            if calc_type == 'selected_floor' and getattr(m, 'selected_floor_id', None):
                floor_id = m.selected_floor_id
                floor = floors_dict.get(floor_id)
                floor_area = floor.area if floor else 0
            else:
                floor_area = total_area
            base_qty = floor_area * (m.factor or 0)
        
        # Handle tile calculation
        tile_width = getattr(m, 'tile_width', 0) or 0
        tile_height = getattr(m, 'tile_height', 0) or 0
        if tile_width > 0 and tile_height > 0 and floor_area > 0:
            tile_area_m2 = (tile_width / 100) * (tile_height / 100)
            if tile_area_m2 > 0:
                base_qty = floor_area / tile_area_m2
        
        # Apply waste
        waste_pct = getattr(m, 'waste_percentage', 0) or 0
        final_qty = base_qty * (1 + waste_pct / 100)
        
        # Skip items with zero quantity
        if final_qty <= 0:
            continue
        
        mat_data = {
            'item_name': m.item_name,
            'unit': m.unit,
            'factor': m.factor or 0,
            'waste_percentage': waste_pct,
            'quantity': round(final_qty, 2),
            'unit_price': m.unit_price or 0,
            'total_price': round(final_qty * (m.unit_price or 0), 2),
            'calculation_method': calc_method
        }
        
        if calc_type == 'selected_floor' and getattr(m, 'selected_floor_id', None):
            floor_id = m.selected_floor_id
            if floor_id not in materials_by_floor:
                materials_by_floor[floor_id] = []
            materials_by_floor[floor_id].append(mat_data)
        else:
            materials_all_floors.append(mat_data)
    
    # Headers for materials table
    headers = ["المادة", "الوحدة", "المعامل", "الهالك%", "الكمية", "السعر", "الإجمالي", "ملاحظات"]
    
    # === Section: Materials for All Floors ===
    if materials_all_floors:
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row=row, column=1, value="▸ مواد جميع الأدوار")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = subheader_fill
        cell.alignment = center_align
        row += 1
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        row += 1
        
        # Write data
        section_total = 0
        for mat in materials_all_floors:
            ws.cell(row=row, column=1, value=mat['item_name']).border = border
            ws.cell(row=row, column=2, value=mat['unit']).border = border
            ws.cell(row=row, column=3, value=mat['factor']).border = border
            ws.cell(row=row, column=4, value=mat['waste_percentage']).border = border
            ws.cell(row=row, column=5, value=mat['quantity']).border = border
            ws.cell(row=row, column=6, value=mat['unit_price']).border = border
            ws.cell(row=row, column=7, value=mat['total_price']).border = border
            ws.cell(row=row, column=8, value="جميع الأدوار").border = border
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = data_fill
            section_total += mat['total_price']
            row += 1
        
        # Section total
        ws.cell(row=row, column=6, value="إجمالي القسم:").font = Font(bold=True)
        ws.cell(row=row, column=7, value=section_total).font = Font(bold=True)
        ws.cell(row=row, column=7).fill = total_fill
        row += 2
    
    # === Sections: Materials by Floor ===
    for floor in floors:
        floor_id = str(floor.id)
        if floor_id not in materials_by_floor:
            continue
        
        floor_materials = materials_by_floor[floor_id]
        if not floor_materials:
            continue
        
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row=row, column=1, value=f"▸ {floor.floor_name} (مساحة: {floor.area} م²)")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = subheader_fill
        cell.alignment = center_align
        row += 1
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        row += 1
        
        # Write data
        section_total = 0
        for mat in floor_materials:
            ws.cell(row=row, column=1, value=mat['item_name']).border = border
            ws.cell(row=row, column=2, value=mat['unit']).border = border
            ws.cell(row=row, column=3, value=mat['factor']).border = border
            ws.cell(row=row, column=4, value=mat['waste_percentage']).border = border
            ws.cell(row=row, column=5, value=mat['quantity']).border = border
            ws.cell(row=row, column=6, value=mat['unit_price']).border = border
            ws.cell(row=row, column=7, value=mat['total_price']).border = border
            ws.cell(row=row, column=8, value=floor.floor_name).border = border
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = data_fill
            section_total += mat['total_price']
            row += 1
        
        # Section total
        ws.cell(row=row, column=6, value="إجمالي الدور:").font = Font(bold=True)
        ws.cell(row=row, column=7, value=section_total).font = Font(bold=True)
        ws.cell(row=row, column=7).fill = total_fill
        row += 2
    
    # Grand Total
    grand_total = sum(m['total_price'] for m in materials_all_floors) + sum(
        sum(m['total_price'] for m in mats) for mats in materials_by_floor.values()
    )
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="الإجمالي الكلي:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).alignment = right_align
    ws.cell(row=row, column=7, value=grand_total).font = Font(bold=True, size=12)
    ws.cell(row=row, column=7).fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    ws.cell(row=row, column=7).font = Font(bold=True, size=12, color="FFFFFF")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    
    # ==================== Sheet 2: ملخص حسب الدور ====================
    ws2 = wb.create_sheet("ملخص حسب الدور")
    ws2.sheet_view.rightToLeft = True
    
    # Title
    ws2.merge_cells('A1:E1')
    ws2['A1'] = f"ملخص طلبات المواد حسب الدور - {project.name}"
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = center_align
    
    # Headers
    summary_headers = ["الدور", "المساحة (م²)", "عدد المواد", "إجمالي القيمة", "النسبة %"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    row = 4
    
    # All floors summary
    if materials_all_floors:
        all_floors_total = sum(m['total_price'] for m in materials_all_floors)
        ws2.cell(row=row, column=1, value="جميع الأدوار").border = border
        ws2.cell(row=row, column=2, value=total_area).border = border
        ws2.cell(row=row, column=3, value=len(materials_all_floors)).border = border
        ws2.cell(row=row, column=4, value=all_floors_total).border = border
        ws2.cell(row=row, column=5, value=f"{(all_floors_total/grand_total*100):.1f}%" if grand_total > 0 else "0%").border = border
        row += 1
    
    # Per floor summary
    for floor in floors:
        floor_id = str(floor.id)
        if floor_id in materials_by_floor:
            floor_materials = materials_by_floor[floor_id]
            floor_total = sum(m['total_price'] for m in floor_materials)
            ws2.cell(row=row, column=1, value=floor.floor_name).border = border
            ws2.cell(row=row, column=2, value=floor.area).border = border
            ws2.cell(row=row, column=3, value=len(floor_materials)).border = border
            ws2.cell(row=row, column=4, value=floor_total).border = border
            ws2.cell(row=row, column=5, value=f"{(floor_total/grand_total*100):.1f}%" if grand_total > 0 else "0%").border = border
            row += 1
    
    # Total row
    ws2.cell(row=row, column=1, value="الإجمالي").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=total_area).font = Font(bold=True)
    total_items_count = len(materials_all_floors) + sum(len(m) for m in materials_by_floor.values())
    ws2.cell(row=row, column=3, value=total_items_count).font = Font(bold=True)
    ws2.cell(row=row, column=4, value=grand_total).font = Font(bold=True)
    ws2.cell(row=row, column=5, value="100%").font = Font(bold=True)
    for c in range(1, 6):
        ws2.cell(row=row, column=c).fill = total_fill
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 12
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from urllib.parse import quote
    safe_filename = f"Materials_Requests_{project_id[:8]}.xlsx"
    encoded_filename = quote(f"طلبات_المواد_{project.name}.xlsx")
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}"
        }
    )
