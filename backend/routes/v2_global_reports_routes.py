"""
V2 Global Reports Routes - تقارير شاملة لكل النظام
يشمل: تقارير المباني، أوامر الشراء، التوريد والاستلام
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, desc
from datetime import datetime, timezone
import io

from database import (
    get_postgres_session, 
    Project, PurchaseOrder, PurchaseOrderItem, 
    ProjectAreaMaterial, PriceCatalogItem, DeliveryRecord,
    Supplier
)
from routes.v2_auth_routes import get_current_user, UserRole

router = APIRouter(
    prefix="/api/v2/reports",
    tags=["V2 Global Reports"]
)


def require_reports_access(user):
    """Check if user has access to reports"""
    allowed_roles = [
        UserRole.QUANTITY_ENGINEER,
        UserRole.PROCUREMENT_MANAGER,
        UserRole.GENERAL_MANAGER,
        UserRole.SYSTEM_ADMIN
    ]
    if user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")


# ==================== تقارير شاملة ====================

@router.get("/global-summary")
async def get_global_summary(
    project_id: Optional[str] = None,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    تقرير شامل يجمع كل البيانات:
    - ملخص المباني والكميات
    - ملخص أوامر الشراء
    - ملخص التوريد والاستلام
    """
    require_reports_access(current_user)
    
    # ========== 1. تقارير المباني (كميات المناطق) ==========
    buildings_query = select(ProjectAreaMaterial)
    if project_id:
        buildings_query = buildings_query.where(ProjectAreaMaterial.project_id == project_id)
    
    buildings_result = await session.execute(buildings_query)
    area_materials = buildings_result.scalars().all()
    
    # Get projects with floors for calculating quantities
    projects_query = select(Project)
    projects_result = await session.execute(projects_query)
    all_projects = {p.id: p for p in projects_result.scalars().all()}
    
    # جلب الأدوار لكل مشروع
    from database import ProjectFloor
    floors_query = select(ProjectFloor)
    floors_result = await session.execute(floors_query)
    all_floors = floors_result.scalars().all()
    floors_by_project = {}
    for f in all_floors:
        if f.project_id not in floors_by_project:
            floors_by_project[f.project_id] = []
        floors_by_project[f.project_id].append(f)
    
    # حساب الكمية لكل مادة
    def calculate_material_quantity(m, project_floors):
        """حساب كمية المادة بناءً على الأدوار"""
        if not project_floors:
            return m.direct_quantity or 0
        
        # حساب مساحة الأدوار
        if m.calculation_type == "selected_floor" and m.selected_floor_id:
            floor = next((f for f in project_floors if f.id == m.selected_floor_id), None)
            floor_area = floor.area if floor else 0
        else:
            floor_area = sum(f.area or 0 for f in project_floors)
        
        if m.calculation_method == "direct":
            return m.direct_quantity or 0
        else:
            # factor method
            raw_qty = floor_area * (m.factor or 0)
            # تحويل للطن إذا كانت الوحدة طن
            if m.unit and 'طن' in m.unit:
                return raw_qty / 1000
            return raw_qty
    
    # حساب إجماليات المباني
    total_buildings_qty = 0
    total_buildings_value = 0
    buildings_by_project = {}
    
    for m in area_materials:
        project_floors = floors_by_project.get(m.project_id, [])
        qty = calculate_material_quantity(m, project_floors)
        value = qty * (m.unit_price or 0)
        
        total_buildings_qty += qty
        total_buildings_value += value
        
        proj_name = all_projects.get(m.project_id, {})
        proj_name = getattr(proj_name, 'name', 'غير محدد') if proj_name else "غير محدد"
        
        if proj_name not in buildings_by_project:
            buildings_by_project[proj_name] = {
                "project_id": m.project_id,
                "project_name": proj_name,
                "items_count": 0,
                "total_quantity": 0,
                "total_value": 0,
                "items": []
            }
        buildings_by_project[proj_name]["items_count"] += 1
        buildings_by_project[proj_name]["total_quantity"] += qty
        buildings_by_project[proj_name]["total_value"] += value
        buildings_by_project[proj_name]["items"].append({
            "item_name": m.item_name,
            "unit": m.unit,
            "quantity": round(qty, 2),
            "unit_price": m.unit_price,
            "total_price": round(value, 2)
        })
    
    # ========== 2. تقارير أوامر الشراء ==========
    orders_query = select(PurchaseOrder)
    if project_id:
        orders_query = orders_query.where(PurchaseOrder.project_id == project_id)
    
    orders_result = await session.execute(orders_query)
    orders = orders_result.scalars().all()
    
    # حساب إجماليات الأوامر
    total_orders = len(orders)
    total_orders_value = sum(o.total_amount or 0 for o in orders)
    
    # تصنيف حسب الحالة
    orders_by_status = {}
    for o in orders:
        status = o.status or "unknown"
        if status not in orders_by_status:
            orders_by_status[status] = {"count": 0, "value": 0}
        orders_by_status[status]["count"] += 1
        orders_by_status[status]["value"] += o.total_amount or 0
    
    # تجميع حسب المورد
    orders_by_supplier = {}
    for o in orders:
        supplier = o.supplier_name or "غير محدد"
        if supplier not in orders_by_supplier:
            orders_by_supplier[supplier] = {"count": 0, "value": 0}
        orders_by_supplier[supplier]["count"] += 1
        orders_by_supplier[supplier]["value"] += o.total_amount or 0
    
    # أوامر حسب المشروع
    orders_by_project = {}
    for o in orders:
        proj_name = o.project_name or "غير محدد"
        if proj_name not in orders_by_project:
            orders_by_project[proj_name] = {
                "project_name": proj_name,
                "total_orders": 0,
                "total_value": 0,
                "delivered": 0,
                "pending": 0
            }
        orders_by_project[proj_name]["total_orders"] += 1
        orders_by_project[proj_name]["total_value"] += o.total_amount or 0
        if o.status == "delivered":
            orders_by_project[proj_name]["delivered"] += 1
        elif o.status in ["pending", "approved", "pending_gm_approval", "pending_procurement_confirmation"]:
            orders_by_project[proj_name]["pending"] += 1
    
    # ========== 3. تقارير التوريد والاستلام ==========
    # جلب عناصر أوامر الشراء مع الكميات المستلمة
    items_query = select(PurchaseOrderItem)
    items_result = await session.execute(items_query)
    all_items = items_result.scalars().all()
    
    # فلترة حسب المشروع إذا تم تحديده
    if project_id:
        order_ids = [o.id for o in orders]
        all_items = [i for i in all_items if i.order_id in order_ids]
    
    total_ordered_qty = sum(i.quantity or 0 for i in all_items)
    total_received_qty = sum(i.received_quantity or 0 for i in all_items)
    total_remaining_qty = total_ordered_qty - total_received_qty
    
    # الأصناف المتأخرة (مطلوبة ولم تُستلم بالكامل)
    pending_items = []
    for item in all_items:
        remaining = (item.quantity or 0) - (item.received_quantity or 0)
        if remaining > 0:
            pending_items.append({
                "item_name": item.name,
                "unit": item.unit,
                "ordered_qty": item.quantity,
                "received_qty": item.received_quantity or 0,
                "remaining_qty": remaining,
                "order_id": item.order_id
            })
    
    # تجميع التوريد حسب المشروع
    supply_by_project = {}
    for o in orders:
        proj_name = o.project_name or "غير محدد"
        if proj_name not in supply_by_project:
            supply_by_project[proj_name] = {
                "project_name": proj_name,
                "ordered_qty": 0,
                "received_qty": 0,
                "remaining_qty": 0,
                "completion_rate": 0
            }
    
    for item in all_items:
        # البحث عن المشروع
        order = next((o for o in orders if o.id == item.order_id), None)
        if order:
            proj_name = order.project_name or "غير محدد"
            if proj_name in supply_by_project:
                supply_by_project[proj_name]["ordered_qty"] += item.quantity or 0
                supply_by_project[proj_name]["received_qty"] += item.received_quantity or 0
    
    # حساب المتبقي ونسبة الإنجاز
    for proj_name, data in supply_by_project.items():
        data["remaining_qty"] = data["ordered_qty"] - data["received_qty"]
        data["completion_rate"] = round((data["received_qty"] / data["ordered_qty"] * 100), 1) if data["ordered_qty"] > 0 else 0
    
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "project_filter": project_id,
        
        # ملخص عام
        "overview": {
            "total_projects": len(set(o.project_id for o in orders if o.project_id)),
            "total_orders": total_orders,
            "total_orders_value": round(total_orders_value, 2),
            "total_buildings_items": len(area_materials),
            "total_buildings_value": round(total_buildings_value, 2),
            "overall_delivery_rate": round((total_received_qty / total_ordered_qty * 100), 1) if total_ordered_qty > 0 else 0
        },
        
        # تقارير المباني
        "buildings": {
            "total_items": len(area_materials),
            "total_quantity": round(total_buildings_qty, 2),
            "total_value": round(total_buildings_value, 2),
            "by_project": list(buildings_by_project.values())
        },
        
        # تقارير أوامر الشراء
        "purchase_orders": {
            "total_orders": total_orders,
            "total_value": round(total_orders_value, 2),
            "by_status": orders_by_status,
            "by_supplier": dict(sorted(orders_by_supplier.items(), key=lambda x: x[1]["value"], reverse=True)[:10]),
            "by_project": list(orders_by_project.values())
        },
        
        # تقارير التوريد
        "supply": {
            "total_ordered_qty": round(total_ordered_qty, 2),
            "total_received_qty": round(total_received_qty, 2),
            "total_remaining_qty": round(total_remaining_qty, 2),
            "completion_rate": round((total_received_qty / total_ordered_qty * 100), 1) if total_ordered_qty > 0 else 0,
            "pending_items_count": len(pending_items),
            "pending_items": pending_items[:20],  # أول 20 صنف متأخر
            "by_project": list(supply_by_project.values())
        }
    }


@router.get("/buildings-summary")
async def get_buildings_summary(
    project_id: Optional[str] = None,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير تفصيلي لكميات المباني"""
    require_reports_access(current_user)
    
    query = select(ProjectAreaMaterial)
    if project_id:
        query = query.where(ProjectAreaMaterial.project_id == project_id)
    
    result = await session.execute(query)
    materials = result.scalars().all()
    
    # Get projects
    projects_query = select(Project)
    projects_result = await session.execute(projects_query)
    all_projects = {p.id: p for p in projects_result.scalars().all()}
    
    # Get floors
    from database import ProjectFloor
    floors_query = select(ProjectFloor)
    floors_result = await session.execute(floors_query)
    all_floors = floors_result.scalars().all()
    floors_by_project = {}
    for f in all_floors:
        if f.project_id not in floors_by_project:
            floors_by_project[f.project_id] = []
        floors_by_project[f.project_id].append(f)
    
    def calc_qty(m, project_floors):
        if not project_floors:
            return m.direct_quantity or 0
        if m.calculation_type == "selected_floor" and m.selected_floor_id:
            floor = next((f for f in project_floors if f.id == m.selected_floor_id), None)
            floor_area = floor.area if floor else 0
        else:
            floor_area = sum(f.area or 0 for f in project_floors)
        if m.calculation_method == "direct":
            return m.direct_quantity or 0
        raw_qty = floor_area * (m.factor or 0)
        if m.unit and 'طن' in m.unit:
            return raw_qty / 1000
        return raw_qty
    
    # تجميع حسب المشروع
    by_project = {}
    total_value = 0
    
    for m in materials:
        project = all_projects.get(m.project_id)
        proj_name = project.name if project else "غير محدد"
        project_floors = floors_by_project.get(m.project_id, [])
        qty = calc_qty(m, project_floors)
        value = qty * (m.unit_price or 0)
        total_value += value
        
        if proj_name not in by_project:
            by_project[proj_name] = {
                "project_id": m.project_id,
                "project_name": proj_name,
                "total_items": 0,
                "total_value": 0,
                "items": []
            }
        by_project[proj_name]["total_items"] += 1
        by_project[proj_name]["total_value"] += value
        by_project[proj_name]["items"].append({
            "id": m.id,
            "item_name": m.item_name,
            "unit": m.unit,
            "quantity": round(qty, 2),
            "unit_price": m.unit_price,
            "total_price": round(value, 2),
            "calculation_type": m.calculation_type,
            "notes": m.notes
        })
    
    return {
        "total_items": len(materials),
        "total_value": round(total_value, 2),
        "by_project": list(by_project.values())
    }


@router.get("/orders-summary")
async def get_orders_summary(
    project_id: Optional[str] = None,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير تفصيلي لأوامر الشراء"""
    require_reports_access(current_user)
    
    query = select(PurchaseOrder)
    if project_id:
        query = query.where(PurchaseOrder.project_id == project_id)
    query = query.order_by(desc(PurchaseOrder.created_at))
    
    result = await session.execute(query)
    orders = result.scalars().all()
    
    # تجميع حسب الحالة
    by_status = {}
    for o in orders:
        status = o.status or "unknown"
        status_ar = {
            "pending": "قيد الانتظار",
            "approved": "معتمد",
            "delivered": "تم التسليم",
            "pending_gm_approval": "بانتظار موافقة المدير العام",
            "pending_procurement_confirmation": "بانتظار تأكيد المشتريات",
            "rejected": "مرفوض",
            "cancelled": "ملغي"
        }.get(status, status)
        
        if status_ar not in by_status:
            by_status[status_ar] = {"count": 0, "value": 0, "orders": []}
        by_status[status_ar]["count"] += 1
        by_status[status_ar]["value"] += o.total_amount or 0
        by_status[status_ar]["orders"].append({
            "id": o.id,
            "order_number": o.order_number,
            "project_name": o.project_name,
            "supplier_name": o.supplier_name,
            "total_amount": o.total_amount,
            "created_at": o.created_at.isoformat() if o.created_at else None
        })
    
    # تجميع حسب المورد
    by_supplier = {}
    for o in orders:
        supplier = o.supplier_name or "غير محدد"
        if supplier not in by_supplier:
            by_supplier[supplier] = {"count": 0, "value": 0}
        by_supplier[supplier]["count"] += 1
        by_supplier[supplier]["value"] += o.total_amount or 0
    
    return {
        "total_orders": len(orders),
        "total_value": round(sum(o.total_amount or 0 for o in orders), 2),
        "by_status": by_status,
        "by_supplier": dict(sorted(by_supplier.items(), key=lambda x: x[1]["value"], reverse=True))
    }


@router.get("/supply-summary")
async def get_supply_summary(
    project_id: Optional[str] = None,
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير تفصيلي للتوريد والاستلام"""
    require_reports_access(current_user)
    
    # جلب الأوامر
    orders_query = select(PurchaseOrder)
    if project_id:
        orders_query = orders_query.where(PurchaseOrder.project_id == project_id)
    
    orders_result = await session.execute(orders_query)
    orders = orders_result.scalars().all()
    order_ids = [o.id for o in orders]
    
    # جلب العناصر
    items_query = select(PurchaseOrderItem)
    if order_ids:
        items_query = items_query.where(PurchaseOrderItem.order_id.in_(order_ids))
    
    items_result = await session.execute(items_query)
    items = items_result.scalars().all()
    
    # تجميع البيانات
    fully_received = []
    partially_received = []
    not_received = []
    
    for item in items:
        ordered = item.quantity or 0
        received = item.received_quantity or 0
        remaining = ordered - received
        
        # البحث عن الأمر
        order = next((o for o in orders if o.id == item.order_id), None)
        
        item_data = {
            "item_name": item.name,
            "unit": item.unit,
            "ordered_qty": ordered,
            "received_qty": received,
            "remaining_qty": remaining,
            "completion_rate": round((received / ordered * 100), 1) if ordered > 0 else 0,
            "order_number": order.order_number if order else None,
            "project_name": order.project_name if order else None,
            "supplier_name": order.supplier_name if order else None
        }
        
        if remaining == 0 and ordered > 0:
            fully_received.append(item_data)
        elif received > 0:
            partially_received.append(item_data)
        elif ordered > 0:
            not_received.append(item_data)
    
    total_ordered = sum(i.quantity or 0 for i in items)
    total_received = sum(i.received_quantity or 0 for i in items)
    
    return {
        "summary": {
            "total_ordered_qty": round(total_ordered, 2),
            "total_received_qty": round(total_received, 2),
            "total_remaining_qty": round(total_ordered - total_received, 2),
            "completion_rate": round((total_received / total_ordered * 100), 1) if total_ordered > 0 else 0
        },
        "fully_received": {
            "count": len(fully_received),
            "items": fully_received[:50]
        },
        "partially_received": {
            "count": len(partially_received),
            "items": partially_received[:50]
        },
        "not_received": {
            "count": len(not_received),
            "items": not_received[:50]
        }
    }


@router.get("/export/excel")
async def export_global_report_excel(
    project_id: Optional[str] = None,
    report_type: str = Query("all", description="نوع التقرير: all, buildings, orders, supply"),
    current_user = Depends(get_current_user),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير التقارير إلى Excel"""
    require_reports_access(current_user)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    title_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ========== ورقة الملخص ==========
    ws_summary = wb.active
    ws_summary.title = "ملخص شامل"
    ws_summary.sheet_view.rightToLeft = True
    
    # جلب البيانات
    # المباني
    buildings_query = select(ProjectAreaMaterial)
    if project_id:
        buildings_query = buildings_query.where(ProjectAreaMaterial.project_id == project_id)
    buildings_result = await session.execute(buildings_query)
    area_materials = buildings_result.scalars().all()
    
    # الأوامر
    orders_query = select(PurchaseOrder)
    if project_id:
        orders_query = orders_query.where(PurchaseOrder.project_id == project_id)
    orders_result = await session.execute(orders_query)
    orders = orders_result.scalars().all()
    
    # العناصر
    items_query = select(PurchaseOrderItem)
    items_result = await session.execute(items_query)
    all_items = items_result.scalars().all()
    if project_id:
        order_ids = [o.id for o in orders]
        all_items = [i for i in all_items if i.order_id in order_ids]
    
    # كتابة الملخص
    row = 1
    ws_summary.merge_cells(f'A{row}:D{row}')
    ws_summary[f'A{row}'] = "تقرير شامل - نظام إدارة طلبات المواد"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary[f'A{row}'].fill = title_fill
    
    row += 2
    ws_summary[f'A{row}'] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    row += 2
    ws_summary[f'A{row}'] = "ملخص المباني والكميات"
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_summary[f'A{row}'] = f"إجمالي الأصناف: {len(area_materials)}"
    row += 1
    total_buildings_value = sum((m.calculated_quantity or 0) * (m.unit_price or 0) for m in area_materials)
    ws_summary[f'A{row}'] = f"إجمالي القيمة: {total_buildings_value:,.2f} ريال"
    
    row += 2
    ws_summary[f'A{row}'] = "ملخص أوامر الشراء"
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_summary[f'A{row}'] = f"إجمالي الأوامر: {len(orders)}"
    row += 1
    total_orders_value = sum(o.total_amount or 0 for o in orders)
    ws_summary[f'A{row}'] = f"إجمالي القيمة: {total_orders_value:,.2f} ريال"
    
    row += 2
    ws_summary[f'A{row}'] = "ملخص التوريد والاستلام"
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    total_ordered = sum(i.quantity or 0 for i in all_items)
    total_received = sum(i.received_quantity or 0 for i in all_items)
    ws_summary[f'A{row}'] = f"الكميات المطلوبة: {total_ordered:,.2f}"
    row += 1
    ws_summary[f'A{row}'] = f"الكميات المستلمة: {total_received:,.2f}"
    row += 1
    ws_summary[f'A{row}'] = f"الكميات المتبقية: {(total_ordered - total_received):,.2f}"
    row += 1
    completion_rate = round((total_received / total_ordered * 100), 1) if total_ordered > 0 else 0
    ws_summary[f'A{row}'] = f"نسبة الإنجاز: {completion_rate}%"
    
    # ========== ورقة المباني ==========
    if report_type in ["all", "buildings"]:
        ws_buildings = wb.create_sheet("كميات المباني")
        ws_buildings.sheet_view.rightToLeft = True
        
        headers = ['المشروع', 'اسم الصنف', 'الوحدة', 'الكمية', 'سعر الوحدة', 'الإجمالي']
        for col, header in enumerate(headers, 1):
            cell = ws_buildings.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_num, m in enumerate(area_materials, 2):
            ws_buildings.cell(row=row_num, column=1, value=m.project_name).border = thin_border
            ws_buildings.cell(row=row_num, column=2, value=m.item_name).border = thin_border
            ws_buildings.cell(row=row_num, column=3, value=m.unit).border = thin_border
            ws_buildings.cell(row=row_num, column=4, value=m.calculated_quantity).border = thin_border
            ws_buildings.cell(row=row_num, column=5, value=m.unit_price).border = thin_border
            ws_buildings.cell(row=row_num, column=6, value=round((m.calculated_quantity or 0) * (m.unit_price or 0), 2)).border = thin_border
        
        ws_buildings.column_dimensions['A'].width = 25
        ws_buildings.column_dimensions['B'].width = 30
        ws_buildings.column_dimensions['C'].width = 12
        ws_buildings.column_dimensions['D'].width = 15
        ws_buildings.column_dimensions['E'].width = 15
        ws_buildings.column_dimensions['F'].width = 15
    
    # ========== ورقة أوامر الشراء ==========
    if report_type in ["all", "orders"]:
        ws_orders = wb.create_sheet("أوامر الشراء")
        ws_orders.sheet_view.rightToLeft = True
        
        headers = ['رقم الأمر', 'المشروع', 'المورد', 'الحالة', 'المبلغ', 'التاريخ']
        for col, header in enumerate(headers, 1):
            cell = ws_orders.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        status_ar = {
            "pending": "قيد الانتظار",
            "approved": "معتمد",
            "delivered": "تم التسليم",
            "pending_gm_approval": "بانتظار المدير العام",
            "pending_procurement_confirmation": "بانتظار المشتريات",
            "rejected": "مرفوض"
        }
        
        for row_num, o in enumerate(orders, 2):
            ws_orders.cell(row=row_num, column=1, value=o.order_number).border = thin_border
            ws_orders.cell(row=row_num, column=2, value=o.project_name).border = thin_border
            ws_orders.cell(row=row_num, column=3, value=o.supplier_name).border = thin_border
            ws_orders.cell(row=row_num, column=4, value=status_ar.get(o.status, o.status)).border = thin_border
            ws_orders.cell(row=row_num, column=5, value=o.total_amount).border = thin_border
            ws_orders.cell(row=row_num, column=6, value=o.created_at.strftime('%Y-%m-%d') if o.created_at else '').border = thin_border
        
        ws_orders.column_dimensions['A'].width = 15
        ws_orders.column_dimensions['B'].width = 25
        ws_orders.column_dimensions['C'].width = 25
        ws_orders.column_dimensions['D'].width = 20
        ws_orders.column_dimensions['E'].width = 15
        ws_orders.column_dimensions['F'].width = 15
    
    # ========== ورقة التوريد ==========
    if report_type in ["all", "supply"]:
        ws_supply = wb.create_sheet("التوريد والاستلام")
        ws_supply.sheet_view.rightToLeft = True
        
        headers = ['الصنف', 'الوحدة', 'المطلوب', 'المستلم', 'المتبقي', 'نسبة الإنجاز', 'رقم الأمر', 'المورد']
        for col, header in enumerate(headers, 1):
            cell = ws_supply.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_num, item in enumerate(all_items, 2):
            order = next((o for o in orders if o.id == item.order_id), None)
            ordered = item.quantity or 0
            received = item.received_quantity or 0
            remaining = ordered - received
            rate = round((received / ordered * 100), 1) if ordered > 0 else 0
            
            ws_supply.cell(row=row_num, column=1, value=item.name).border = thin_border
            ws_supply.cell(row=row_num, column=2, value=item.unit).border = thin_border
            ws_supply.cell(row=row_num, column=3, value=ordered).border = thin_border
            ws_supply.cell(row=row_num, column=4, value=received).border = thin_border
            ws_supply.cell(row=row_num, column=5, value=remaining).border = thin_border
            ws_supply.cell(row=row_num, column=6, value=f"{rate}%").border = thin_border
            ws_supply.cell(row=row_num, column=7, value=order.order_number if order else '').border = thin_border
            ws_supply.cell(row=row_num, column=8, value=order.supplier_name if order else '').border = thin_border
        
        ws_supply.column_dimensions['A'].width = 30
        ws_supply.column_dimensions['B'].width = 12
        ws_supply.column_dimensions['C'].width = 12
        ws_supply.column_dimensions['D'].width = 12
        ws_supply.column_dimensions['E'].width = 12
        ws_supply.column_dimensions['F'].width = 15
        ws_supply.column_dimensions['G'].width = 15
        ws_supply.column_dimensions['H'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=global_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"}
    )
